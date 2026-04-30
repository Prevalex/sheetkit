from pathlib import Path
import json
import shutil
import uuid
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import sheetkit.themes as themes_module
from sheetkit.themes import (
    extract_formatter_from_sheet,
    extract_formatter_range_from_sheet,
    extract_formatter_range_to_file,
    extract_formatter_to_file,
    get_theme,
    list_preset_formatters,
    list_preset_themes,
    load_formatter,
    load_format_preset,
    load_preset_file,
    save_formatter,
)
from sheetkit.tools import export_excel_theme


@pytest.fixture
def workspace_tmp_dir() -> None:
    base_dir = Path(__file__).resolve().parent / "_tmp"
    temp_dir = base_dir / f"presets-{uuid.uuid4().hex}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        yield temp_dir
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture(autouse=True)
def restore_themes_cache() -> None:
    saved = {key: dict(value) for key, value in themes_module.THEMES.items()}
    try:
        yield
    finally:
        themes_module.THEMES.clear()
        themes_module.THEMES.update(saved)


def _write_test_thmx(path: Path) -> None:
    theme_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Preset Export Theme">
  <a:themeElements>
    <a:clrScheme name="Preset Colors">
      <a:dk1><a:srgbClr val="111111"/></a:dk1>
      <a:lt1><a:srgbClr val="FFFFFF"/></a:lt1>
      <a:dk2><a:srgbClr val="222222"/></a:dk2>
      <a:lt2><a:srgbClr val="EEEEEE"/></a:lt2>
      <a:accent1><a:srgbClr val="4472C4"/></a:accent1>
      <a:accent2><a:srgbClr val="ED7D31"/></a:accent2>
      <a:hlink><a:srgbClr val="0563C1"/></a:hlink>
      <a:folHlink><a:srgbClr val="954F72"/></a:folHlink>
    </a:clrScheme>
    <a:fontScheme name="Preset Fonts">
      <a:majorFont><a:latin typeface="Aptos Display"/></a:majorFont>
      <a:minorFont><a:latin typeface="Aptos"/></a:minorFont>
    </a:fontScheme>
  </a:themeElements>
</a:theme>
"""
    with ZipFile(path, "w") as archive:
        archive.writestr("theme/theme/theme1.xml", theme_xml)


def test_get_theme_returns_built_in_json_theme() -> None:
    theme = get_theme("office_theme")
    assert theme["name"] == "office_theme"
    assert theme["scheme_name"] == "Office Theme"
    assert theme["colors"]["accent1"] == "4472C4"


def test_load_preset_file_loads_theme_json_file(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "custom_theme.json"
    file_name.write_text(
        json.dumps(
            {
                "kind": "themes",
                "name": "custom_theme",
                "scheme_name": "Custom Theme",
                "colors": {"accent1": "123456"},
                "fonts": {"minor": "Calibri"},
            }
        ),
        encoding="utf-8",
    )
    preset = load_preset_file(file_name, expected_kind="themes")
    assert preset["name"] == "custom_theme"
    assert preset["colors"]["accent1"] == "123456"


def test_load_formatter_loads_runtime_formatter_json(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "custom_formatter.json"
    file_name.write_text(
        json.dumps(
            {
                "kind": "formatter",
                "name": "custom_formatter",
                "row": {"0": {"bold": True}, "-1": {"border_bottom": 1}},
                "col": {"1": {"num_format": "0.00"}},
            }
        ),
        encoding="utf-8",
    )
    formatter = load_formatter(file_name)
    assert formatter["row"][0]["bold"] is True
    assert formatter["row"][-1]["border_bottom"] == 1
    assert formatter["col"][1]["num_format"] == "0.00"


def test_load_formatter_rejects_missing_runtime_keys(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "bad_formatter.json"
    file_name.write_text(json.dumps({"kind": "formatter", "row": {}}), encoding="utf-8")
    with pytest.raises(ValueError, match="must contain 'row' and 'col'"):
        load_formatter(file_name)


def test_save_formatter_roundtrip(workspace_tmp_dir: Path) -> None:
    formatter = {"row": {0: {"bold": True}, -1: {"border_bottom": 1}}, "col": {0: {"num_format": "@"}}}
    json_file = workspace_tmp_dir / "saved_formatter.json"
    result = save_formatter(formatter, json_file, name="saved_formatter")
    assert result == json_file
    loaded = load_formatter(json_file)
    assert loaded["row"][0]["bold"] is True
    assert loaded["col"][0]["num_format"] == "@"


def test_export_excel_theme_converts_thmx_to_json(workspace_tmp_dir: Path) -> None:
    thmx_file = workspace_tmp_dir / "theme.thmx"
    json_file = workspace_tmp_dir / "theme.json"
    _write_test_thmx(thmx_file)

    result = export_excel_theme(thmx_file, json_file)
    assert result == json_file

    exported = json.loads(json_file.read_text(encoding="utf-8"))
    assert exported["kind"] == "themes"
    assert exported["scheme_name"] == "Preset Export Theme"
    assert exported["colors"]["accent1"] == "4472C4"
    assert exported["fonts"]["minor"] == "Aptos"


def test_formatter_extraction_wrappers_delegate(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "source.xlsx"
    json_file = workspace_tmp_dir / "fmt.json"
    _create_formatter_source_workbook(file_name)

    formatter = extract_formatter_from_sheet(file_name, columns=3, header_rows=2, zebra=True)
    assert formatter["row"][0]["fg_color"] == "4472C4"

    path = extract_formatter_to_file(
        file_name=file_name,
        json_file=json_file,
        columns=3,
        header_rows=2,
        zebra=True,
        name="a",
    )
    assert path == json_file
    loaded = load_formatter(json_file)
    assert loaded["row"][0]["fg_color"] == "4472C4"


def _create_formatter_source_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Styled"

    header1_fill = PatternFill(patternType="solid", fgColor="4472C4")
    header2_fill = PatternFill(patternType="solid", fgColor="D9E2F3")
    alt_fill = PatternFill(patternType="solid", fgColor="F7F9FC")
    white_font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    dark_font = Font(name="Aptos", size=11, bold=True, color="1F1F1F")
    data_font = Font(name="Aptos", size=11, color="222222")
    thin_blue = Side(style="thin", color="D9E2F3")
    thin_gray = Side(style="thin", color="D9D9D9")

    for col_idx in range(1, 5):
        ws.cell(row=1, column=col_idx, value=f"H1-{col_idx}")
        ws.cell(row=1, column=col_idx).fill = header1_fill
        ws.cell(row=1, column=col_idx).font = white_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=col_idx).border = Border(bottom=thin_blue)

        ws.cell(row=2, column=col_idx, value=f"H2-{col_idx}")
        ws.cell(row=2, column=col_idx).fill = header2_fill
        ws.cell(row=2, column=col_idx).font = dark_font
        ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col_idx).border = Border(bottom=thin_blue)

        ws.cell(row=3, column=col_idx, value=10 * col_idx)
        ws.cell(row=3, column=col_idx).font = data_font
        ws.cell(row=3, column=col_idx).border = Border(bottom=thin_gray)

        ws.cell(row=4, column=col_idx, value=20 * col_idx)
        ws.cell(row=4, column=col_idx).font = data_font
        ws.cell(row=4, column=col_idx).fill = alt_fill
        ws.cell(row=4, column=col_idx).border = Border(bottom=thin_gray)

    ws["A3"].number_format = "@"
    ws["B3"].number_format = "#,##0.00"
    ws["C4"].number_format = "0.0%"

    wb.create_sheet("Other")
    wb.save(path)


def _create_formatter_with_no_fill_base_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Styled"

    header1_fill = PatternFill(patternType="solid", fgColor="3B618E")
    header2_fill = PatternFill(patternType="solid", fgColor="D9E1F2")
    alt_fill = PatternFill(patternType="solid", fgColor="F9F9F9")
    white_font = Font(name="Calibri", size=11, bold=True, color="DCE6F2")
    dark_font = Font(name="Calibri", size=11, bold=True, color="000000")
    data_font = Font(name="Calibri", size=11, color="000000")
    thin = Side(style="thin", color="D9D9D9")

    for col_idx in range(1, 4):
        ws.cell(row=1, column=col_idx, value=f"H1-{col_idx}")
        ws.cell(row=1, column=col_idx).fill = header1_fill
        ws.cell(row=1, column=col_idx).font = white_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=2, column=col_idx, value=None)
        ws.cell(row=2, column=col_idx).fill = header2_fill
        ws.cell(row=2, column=col_idx).font = dark_font
        ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=3, column=col_idx, value=f"R3-{col_idx}")
        ws.cell(row=3, column=col_idx).font = data_font
        ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="left", vertical="top")
        ws.cell(row=3, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=4, column=col_idx, value=f"R4-{col_idx}")
        ws.cell(row=4, column=col_idx).fill = alt_fill
        ws.cell(row=4, column=col_idx).font = data_font
        ws.cell(row=4, column=col_idx).alignment = Alignment(horizontal="left", vertical="top")
        ws.cell(row=4, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb.save(path)


def _create_column_priority_source_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Styled"

    header1_fill = PatternFill(patternType="solid", fgColor="4472C4")
    header2_fill = PatternFill(patternType="solid", fgColor="D9E2F3")
    alt_fill = PatternFill(patternType="solid", fgColor="F7F9FC")
    white_font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Aptos", size=11, color="222222")
    thin_gray = Side(style="thin", color="D9D9D9")

    for row_idx in range(1, 5):
        ws.cell(row=row_idx, column=1, value=f"H1-{row_idx}")
        ws.cell(row=row_idx, column=1).fill = header1_fill
        ws.cell(row=row_idx, column=1).font = white_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1).border = Border(bottom=thin_gray)

        ws.cell(row=row_idx, column=2, value=f"H2-{row_idx}")
        ws.cell(row=row_idx, column=2).fill = header2_fill
        ws.cell(row=row_idx, column=2).font = data_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=2).border = Border(bottom=thin_gray)

        ws.cell(row=row_idx, column=3, value=row_idx * 10)
        ws.cell(row=row_idx, column=3).font = data_font
        ws.cell(row=row_idx, column=3).border = Border(bottom=thin_gray)

        ws.cell(row=row_idx, column=4, value=row_idx * 20)
        ws.cell(row=row_idx, column=4).fill = alt_fill
        ws.cell(row=row_idx, column=4).font = data_font
        ws.cell(row=row_idx, column=4).border = Border(bottom=thin_gray)

    ws["C1"].number_format = "@"
    ws["C2"].number_format = "#,##0.00"
    ws["D4"].number_format = "0.0%"
    wb.save(path)


def _create_full_range_row_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    fills = ["4472C4", "70AD47", "FFC000"]
    for row_idx, fill_color in enumerate(fills, start=1):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
            cell.fill = PatternFill(patternType="solid", fgColor=fill_color)
            cell.font = Font(name="Aptos", size=11, color="FFFFFF" if row_idx == 1 else "000000")
        ws.cell(row=row_idx, column=2).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=3).number_format = "0.0%"
    wb.save(path)


def _create_full_range_col_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    fills = ["4472C4", "70AD47", "FFC000"]
    for col_idx, fill_color in enumerate(fills, start=1):
        for row_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
            cell.fill = PatternFill(patternType="solid", fgColor=fill_color)
            cell.font = Font(name="Aptos", size=11, color="FFFFFF" if col_idx == 1 else "000000")
        ws.cell(row=2, column=col_idx).number_format = "#,##0.00"
        ws.cell(row=3, column=col_idx).number_format = "0.0%"
    wb.save(path)


def test_extract_formatter_uses_active_sheet_by_default(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    _create_formatter_source_workbook(file_name)
    formatter = extract_formatter_from_sheet(file_name, columns=3)
    assert formatter["row"][0]["fg_color"] == "4472C4"
    assert formatter["row"][1]["fg_color"] == "D9E2F3"
    assert formatter["row"][-2]["fg_color"] == "F7F9FC"
    assert formatter["col"][0]["num_format"] == "@"
    assert formatter["col"][1]["num_format"] == "#,##0.00"
    assert formatter["col"][2]["num_format"] == "0.0%"


def test_extract_formatter_accepts_explicit_sheet_name(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    _create_formatter_source_workbook(file_name)
    formatter = extract_formatter_from_sheet(file_name, "Styled", columns=2)
    assert formatter["row"][0]["fg_color"] == "4472C4"
    assert formatter["col"][1]["num_format"] == "#,##0.00"


def test_extract_formatter_raises_for_missing_sheet(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    _create_formatter_source_workbook(file_name)
    with pytest.raises(ValueError, match="Worksheet 'Missing' not found"):
        extract_formatter_from_sheet(file_name, "Missing", columns=1)


def test_extract_formatter_to_file_exports_json(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    json_file = workspace_tmp_dir / "imported_formatter.json"
    _create_formatter_source_workbook(file_name)
    result = extract_formatter_to_file(file_name, json_file, columns=3, name="from_excel")
    assert result == json_file
    exported = load_formatter(json_file)
    assert exported["row"][-2]["fg_color"] == "F7F9FC"
    assert exported["col"][2]["num_format"] == "0.0%"


def test_extract_formatter_treats_no_fill_as_no_color(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "no_fill_base.xlsx"
    _create_formatter_with_no_fill_base_workbook(file_name)
    formatter = extract_formatter_from_sheet(file_name, columns=3, header_rows=2, zebra=True)
    assert formatter["row"][0]["fg_color"] == "3B618E"
    assert formatter["row"][1]["fg_color"] == "D9E1F2"
    assert "fg_color" not in formatter["row"][-1]
    assert "bg_color" not in formatter["row"][-1]
    assert formatter["row"][-2]["fg_color"] == "F9F9F9"


def test_extract_formatter_without_zebra_only_imports_base_row(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    _create_formatter_source_workbook(file_name)
    formatter = extract_formatter_from_sheet(file_name, columns=3, zebra=False)
    assert -1 in formatter["row"]
    assert -2 not in formatter["row"]


def test_extract_formatter_supports_column_priority(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "styled.xlsx"
    _create_column_priority_source_workbook(file_name)

    formatter = extract_formatter_from_sheet(
        file_name,
        rows=4,
        header_rows=2,
        zebra=True,
        priority="col",
    )

    assert formatter["priority"] == "col"
    assert formatter["col"][0]["fg_color"] == "4472C4"
    assert formatter["col"][1]["fg_color"] == "D9E2F3"
    assert formatter["col"][-2]["fg_color"] == "F7F9FC"
    assert formatter["row"][0]["num_format"] == "@"
    assert formatter["row"][1]["num_format"] == "#,##0.00"
    assert formatter["row"][3]["num_format"] == "0.0%"


def test_extract_formatter_range_copies_row_styles_and_column_types(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "range_rows.xlsx"
    json_file = workspace_tmp_dir / "range_rows.json"
    _create_full_range_row_workbook(file_name)

    formatter = extract_formatter_range_from_sheet(file_name, rows=3, columns=3, priority="row")

    assert formatter["priority"] == "row"
    assert formatter["row"][0]["fg_color"] == "4472C4"
    assert formatter["row"][1]["fg_color"] == "70AD47"
    assert formatter["row"][-1]["fg_color"] == "FFC000"
    assert 2 not in formatter["row"]
    assert formatter["col"][1]["num_format"] == "#,##0.00"
    assert formatter["col"][2]["num_format"] == "0.0%"

    result = extract_formatter_range_to_file(
        file_name,
        json_file,
        rows=3,
        columns=3,
        priority="row",
        name="full_rows",
    )
    assert result == json_file
    assert load_formatter(json_file)["row"][-1]["fg_color"] == "FFC000"


def test_extract_formatter_range_copies_column_styles_and_row_types(workspace_tmp_dir: Path) -> None:
    file_name = workspace_tmp_dir / "range_cols.xlsx"
    _create_full_range_col_workbook(file_name)

    formatter = extract_formatter_range_from_sheet(file_name, rows=3, columns=3, priority="col")

    assert formatter["priority"] == "col"
    assert formatter["col"][0]["fg_color"] == "4472C4"
    assert formatter["col"][1]["fg_color"] == "70AD47"
    assert formatter["col"][-1]["fg_color"] == "FFC000"
    assert 2 not in formatter["col"]
    assert formatter["row"][1]["num_format"] == "#,##0.00"
    assert formatter["row"][2]["num_format"] == "0.0%"


def test_list_preset_themes_contains_known_presets() -> None:
    names = list_preset_themes()
    assert "office_theme" in names
    assert "wisp" in names
    assert "facet" in names


def test_list_preset_formatters_contains_known_pairs() -> None:
    names = list_preset_formatters()
    assert "office_theme" in names
    assert "wisp" in names
    assert "facet" in names


def test_load_format_preset_loads_builtin_row_and_col() -> None:
    row_formatter = load_format_preset("office_theme", priority="row")
    col_formatter = load_format_preset("office_theme", priority="col")

    assert row_formatter["priority"] == "row"
    assert col_formatter["priority"] == "col"
    assert 0 in row_formatter["row"]
    assert 0 in col_formatter["col"]

