from pathlib import Path
from datetime import date
import shutil
import uuid

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from sheetkit.creator import _get_row_styles, _resolve_engine_mode_and_path, read_sheet, save_formatted_xlsx
from sheetkit.themes import build_formatter_from_theme, get_theme


@pytest.fixture
def workspace_tmp_dir() -> None:
    base_dir = Path(__file__).resolve().parent / "_tmp"
    temp_dir = base_dir / f"creator-{uuid.uuid4().hex}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        yield temp_dir
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_save_formatted_xlsx_creates_file_with_xlsxwriter(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "xlsxwriter_report"
    data = [
        ["Name", "Value"],
        ["Alpha", 1.25],
        ["Beta", 2.5],
    ]
    formatter = build_formatter_from_theme(
        get_theme("office_theme"),
        types=["@", "0.00"],
        header=1,
        zebra=True,
    )

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="xlsxwriter",
        mode="new",
        header=1,
    )

    assert saved_path.exists()
    assert saved_path.suffix == ".xlsx"

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A1"].value == "Name"
    assert ws["B2"].value == 1.25
    assert ws["A1"].font.bold is True
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["B2"].number_format == "0.00"


def test_save_formatted_xlsx_creates_file_with_openpyxl(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "openpyxl_report.xlsx"
    data = [
        ["Name", "Value"],
        ["Gamma", 10],
        ["Delta", 20],
    ]
    formatter = build_formatter_from_theme(
        get_theme("office_theme"),
        types=["@", "0"],
        header=1,
        zebra=True,
    )

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=1,
    )

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb == "FF4472C4"
    assert ws["A2"].value == "Gamma"
    assert ws["B2"].number_format == "0"


def test_engine_auto_uses_xlsxwriter_for_new_workbooks(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "auto_new.xlsx"

    engine, mode, path = _resolve_engine_mode_and_path(
        file_name=output,
        engine="auto",
        mode="auto",
        as_template=False,
    )

    assert engine == "xlsxwriter"
    assert mode == "new"
    assert path == output


def test_engine_auto_uses_openpyxl_for_existing_workbooks(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "auto_existing.xlsx"
    output.write_bytes(b"placeholder")

    engine, mode, path = _resolve_engine_mode_and_path(
        file_name=output,
        engine="auto",
        mode="auto",
        as_template=False,
    )

    assert engine == "openpyxl"
    assert mode == "replace"
    assert path == output


def test_engine_auto_uses_openpyxl_for_templates(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "auto_template"

    engine, mode, path = _resolve_engine_mode_and_path(
        file_name=output,
        engine="auto",
        mode="auto",
        as_template=True,
    )

    assert engine == "openpyxl"
    assert mode == "new"
    assert path.suffix == ".xltx"


def test_mode_new_raises_when_target_file_exists(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "existing.xlsx"
    output.write_bytes(b"existing")

    with pytest.raises(FileExistsError, match="mode 'new'"):
        _resolve_engine_mode_and_path(
            file_name=output,
            engine="auto",
            mode="new",
            as_template=False,
        )


def test_save_formatted_xlsx_updates_existing_sheet_and_clears_old_cells(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "update_report.xlsx"
    formatter = build_formatter_from_theme(
        get_theme("office_theme"),
        types=["@", "0"],
        header=1,
        zebra=False,
    )

    save_formatted_xlsx(
        data_table=[
            ["Name", "Value"],
            ["Old", 1],
            ["Extra", 2],
        ],
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=1,
    )

    save_formatted_xlsx(
        data_table=[
            ["Name", "Value"],
            ["New", 9],
        ],
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="update",
        header=1,
        backup=False,
    )

    wb = load_workbook(output)
    ws = wb["Report"]

    assert ws["A1"].value == "Name"
    assert ws["A2"].value == "New"
    assert ws["B2"].value == 9
    assert ws["A3"].value is None
    assert ws["B3"].value is None


def test_save_formatted_xlsx_patch_updates_only_non_none_cells(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_report.xlsx"
    save_formatted_xlsx(
        data_table=[
            ["Name", "Value"],
            ["Old", 1],
            ["Stay", 2],
        ],
        file_name=output,
        sheet_name="Report",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
        header=1,
    )

    save_formatted_xlsx(
        data_table=[
            [None, None],
            ["New", 9],
            [None, 3],
        ],
        file_name=output,
        sheet_name="Report",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="patch",
    )

    ws = load_workbook(output)["Report"]
    assert ws["A1"].value == "Name"
    assert ws["B1"].value == "Value"
    assert ws["A2"].value == "New"
    assert ws["B2"].value == 9
    assert ws["A3"].value == "Stay"
    assert ws["B3"].value == 3


def test_save_formatted_xlsx_patch_skips_formula_and_locked_on_protected_sheet(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_protected.xlsx"
    save_formatted_xlsx(
        data_table=[["Input", "Formula"], [10, None]],
        file_name=output,
        sheet_name="Calc",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    wb = load_workbook(output)
    ws = wb["Calc"]
    ws["B2"] = "=A2*2"
    ws["A2"].protection = Protection(locked=False)
    ws["B2"].protection = Protection(locked=True)
    ws.protection.sheet = True
    wb.save(output)

    save_formatted_xlsx(
        data_table=[["Input", "Formula"], [20, 999]],
        file_name=output,
        sheet_name="Calc",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="patch",
    )

    ws2 = load_workbook(output)["Calc"]
    assert ws2["A2"].value == 20
    assert ws2["B2"].value == "=A2*2"


def test_save_formatted_xlsx_patch_range_writes_from_range_anchor(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_range_anchor.xlsx"
    save_formatted_xlsx(
        data_table=[[1, 2], [3, 4]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    save_formatted_xlsx(
        data_table=[[11], [22], [33]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="patch",
        patch_range="D2:D10",
    )

    ws = load_workbook(output)["Sheet1"]
    assert ws["D2"].value == 11
    assert ws["D3"].value == 22
    assert ws["D4"].value == 33
    assert ws["A1"].value == 1


def test_save_formatted_xlsx_patch_range_soft_clips_by_default(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_range_soft_clip.xlsx"
    save_formatted_xlsx(
        data_table=[[0]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    save_formatted_xlsx(
        data_table=[[1, 2], [3, 4]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="patch",
        patch_range="D2:E2",
        patch_strict_range=False,
    )

    ws = load_workbook(output)["Sheet1"]
    assert ws["D2"].value == 1
    assert ws["E2"].value == 2
    assert ws["D3"].value is None
    assert ws["E3"].value is None


def test_save_formatted_xlsx_patch_range_strict_raises_when_out_of_bounds(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_range_strict.xlsx"
    save_formatted_xlsx(
        data_table=[[0]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    with pytest.raises(ValueError, match="does not fit patch_range"):
        save_formatted_xlsx(
            data_table=[[1, 2], [3, 4]],
            file_name=output,
            sheet_name="Sheet1",
            formatter={"row": {}, "col": {}},
            engine="openpyxl",
            mode="patch",
            patch_range="D2:E2",
            patch_strict_range=True,
        )


def test_save_formatted_xlsx_patch_range_tuple_zero_based_equivalent_to_a1_b2(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_range_tuple.xlsx"
    save_formatted_xlsx(
        data_table=[[0, 0], [0, 0]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    save_formatted_xlsx(
        data_table=[[11, 12], [21, 22]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="patch",
        patch_range=((0, 0), (1, 1)),
    )

    ws = load_workbook(output)["Sheet1"]
    assert ws["A1"].value == 11
    assert ws["B1"].value == 12
    assert ws["A2"].value == 21
    assert ws["B2"].value == 22


def test_save_formatted_xlsx_patch_range_tuple_invalid_raises(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_range_tuple_invalid.xlsx"
    save_formatted_xlsx(
        data_table=[[0]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    with pytest.raises(ValueError, match="Invalid patch_range"):
        save_formatted_xlsx(
            data_table=[[1]],
            file_name=output,
            sheet_name="Sheet1",
            formatter={"row": {}, "col": {}},
            engine="openpyxl",
            mode="patch",
            patch_range=((1, 1), (0, 0)),
        )


def test_save_formatted_xlsx_auto_mode_with_patch_range_uses_patch_not_replace(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "patch_auto_mode.xlsx"
    save_formatted_xlsx(
        data_table=[
            ["H1", "H2", "H3"],
            [1, 2, 3],
            [4, 5, 6],
        ],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="auto",
        mode="new",
    )

    save_formatted_xlsx(
        data_table=[[200]],
        file_name=output,
        sheet_name="Sheet1",
        formatter={"row": {}, "col": {}},
        engine="auto",
        mode="auto",
        patch_range="B2:B2",
    )

    ws = load_workbook(output)["Sheet1"]
    assert ws["A1"].value == "H1"
    assert ws["B1"].value == "H2"
    assert ws["C1"].value == "H3"
    assert ws["A2"].value == 1
    assert ws["B2"].value == 200
    assert ws["C2"].value == 3
    assert ws["A3"].value == 4
    assert ws["B3"].value == 5
    assert ws["C3"].value == 6


def test_read_sheet_reads_full_used_area(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_full.xlsx"
    save_formatted_xlsx(
        data_table=[["H1", "H2"], [1, 2], [3, 4]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    data = read_sheet(output)
    assert data == [["H1", "H2"], [1, 2], [3, 4]]


def test_read_sheet_reads_string_range(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_string_range.xlsx"
    save_formatted_xlsx(
        data_table=[["H1", "H2", "H3"], [10, 20, 30], [40, 50, 60]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    data = read_sheet(output, sheet_range="B2:C3")
    assert data == [[20, 30], [50, 60]]


def test_read_sheet_reads_tuple_range_zero_based(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_tuple_range.xlsx"
    save_formatted_xlsx(
        data_table=[["H1", "H2", "H3"], [10, 20, 30], [40, 50, 60]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    data = read_sheet(output, sheet_range=((1, 1), (2, 2)))
    assert data == [[20, 30], [50, 60]]


def test_read_sheet_data_only_false_returns_formula(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_formula.xlsx"
    save_formatted_xlsx(
        data_table=[["A", "B"], [10, None]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )
    wb = load_workbook(output)
    ws = wb.active
    ws["B2"] = "=A2*2"
    wb.save(output)

    values = read_sheet(output, data_only=False)
    assert values[1][1] == "=A2*2"


def test_read_sheet_data_only_true_returns_calculated_value_or_none(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_formula_data_only.xlsx"
    save_formatted_xlsx(
        data_table=[["A", "B"], [10, None]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )
    wb = load_workbook(output)
    ws = wb.active
    ws["B2"] = "=A2*2"
    wb.save(output)

    values = read_sheet(output, data_only=True)
    # For formula cells without cached result openpyxl returns None in data_only mode.
    assert values[1][1] is None


def test_read_sheet_invalid_tuple_range_raises(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "read_invalid_range.xlsx"
    save_formatted_xlsx(
        data_table=[[1]],
        file_name=output,
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
    )

    with pytest.raises(ValueError, match="Invalid sheet_range"):
        read_sheet(output, sheet_range=((2, 2), (1, 1)))


def test_save_formatted_xlsx_replaces_sheet_and_keeps_backup_copy(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "replace_report.xlsx"
    formatter = build_formatter_from_theme(
        get_theme("office_theme"),
        types=["@", "0.00"],
        header=1,
        zebra=False,
    )

    save_formatted_xlsx(
        data_table=[
            ["Name", "Value"],
            ["Before", 1.5],
        ],
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=1,
    )

    save_formatted_xlsx(
        data_table=[
            ["Name", "Value"],
            ["After", 7.5],
        ],
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="replace",
        header=1,
        backup=True,
    )

    wb = load_workbook(output)

    assert "Report" in wb.sheetnames
    assert "Report (1)" in wb.sheetnames
    assert wb["Report"]["A2"].value == "After"
    assert wb["Report"]["B2"].value == 7.5
    assert wb["Report (1)"]["A2"].value == "Before"
    assert wb["Report (1)"]["B2"].value == 1.5


def test_get_row_styles_starts_zebra_after_explicit_header_rows() -> None:
    style_formatter = {
        0: ["header"],
        -1: ["base"],
        -2: ["alt"],
    }

    assert _get_row_styles(style_formatter, 1) == ["base"]
    assert _get_row_styles(style_formatter, 2) == ["alt"]
    assert _get_row_styles(style_formatter, 3) == ["base"]


def test_save_header_argument_disables_formatter_header_styles(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "header_disabled.xlsx"
    data = [["H1", "H2"], ["A", 1], ["B", 2]]
    formatter = {
        "row": {
            0: {"pattern": "solid", "fg_color": "FF0000"},
            1: {"pattern": "solid", "fg_color": "0000FF"},
            -1: {"pattern": "solid", "fg_color": "FFFFFF"},
            -2: {"pattern": "solid", "fg_color": "EEEEEE"},
        },
        "col": {},
    }

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=0,
    )

    ws = load_workbook(saved_path).active

    assert ws["A1"].fill.fgColor.rgb == "FFFFFFFF"
    assert ws["A2"].fill.fgColor.rgb == "FFEEEEEE"


def test_save_header_argument_ignores_extra_second_header_style(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "single_header.xlsx"
    data = [["H1", "H2"], ["A", 1], ["B", 2]]
    formatter = {
        "row": {
            0: {"pattern": "solid", "fg_color": "FF0000"},
            1: {"pattern": "solid", "fg_color": "0000FF"},
            -1: {"pattern": "solid", "fg_color": "FFFFFF"},
            -2: {"pattern": "solid", "fg_color": "EEEEEE"},
        },
        "col": {},
    }

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=1,
    )

    ws = load_workbook(saved_path).active

    assert ws["A1"].fill.fgColor.rgb == "FFFF0000"
    assert ws["A2"].fill.fgColor.rgb == "FFFFFFFF"
    assert ws["A3"].fill.fgColor.rgb == "FFEEEEEE"


def test_save_header_argument_repeats_available_header_style(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "repeated_header.xlsx"
    data = [["H1", "H2"], ["Sub1", "Sub2"], ["A", 1]]
    formatter = {
        "row": {
            0: {"pattern": "solid", "fg_color": "FF0000"},
            -1: {"pattern": "solid", "fg_color": "FFFFFF"},
        },
        "col": {},
    }

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=2,
    )

    ws = load_workbook(saved_path).active

    assert ws["A1"].fill.fgColor.rgb == "FFFF0000"
    assert ws["A2"].fill.fgColor.rgb == "FFFF0000"
    assert ws["A3"].fill.fgColor.rgb == "FFFFFFFF"


def test_save_without_header_override_keeps_formatter_axis_0_1(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "keep_axis01.xlsx"
    data = [["H1", "H2", "H3"], ["A", 1, 10], ["B", 2, 20]]
    formatter = {
        "priority": "col",
        "col": {
            0: {"pattern": "solid", "fg_color": "FF0000"},
            1: {"pattern": "solid", "fg_color": "0000FF"},
            -1: {"pattern": "solid", "fg_color": "00FF00"},
        },
        "row": {},
    }

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        formatter=formatter,
        engine="openpyxl",
        mode="new",
    )

    ws = load_workbook(saved_path).active

    assert ws["A1"].fill.fgColor.rgb == "FFFF0000"
    assert ws["B1"].fill.fgColor.rgb == "FF0000FF"
    assert ws["C1"].fill.fgColor.rgb == "FF00FF00"


def test_save_formatted_xlsx_infers_num_formats_from_first_data_row(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "inferred_num_formats.xlsx"
    data = [
        ["int", "float", "date", "text", "bool", "none"],
        [1234, 1234.5, date(2026, 4, 23), "abc", True, None],
    ]

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="new",
        header=1,
    )

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A2"].number_format == "#,##0"
    assert ws["B2"].number_format == "#,##0.00"
    assert ws["C2"].number_format == "dd.mm.yyyy"
    assert ws["D2"].number_format == "General"
    assert ws["E2"].number_format == "General"
    assert ws["F2"].number_format == "General"


def test_save_formatted_xlsx_keeps_explicit_num_formats(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "explicit_num_formats.xlsx"
    data = [
        ["int", "float"],
        [10, 3.1415],
    ]

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter={"row": {}, "col": {1: {"num_format": "0.0000"}}},
        engine="openpyxl",
        mode="new",
        header=1,
    )

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A2"].number_format == "#,##0"
    assert ws["B2"].number_format == "0.0000"


def test_save_formatted_xlsx_supports_column_priority(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "column_priority.xlsx"
    data = [
        ["Name", "Alpha", "Beta", "Gamma"],
        ["Value", 1.25, 2.5, 3.75],
    ]
    formatter = build_formatter_from_theme(
        get_theme("office_theme"),
        priority="col",
        header=1,
        zebra=True,
        types=["@", "0.00"],
    )

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine="openpyxl",
        mode="new",
        header=1,
    )

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A1"].fill.fgColor.rgb == "FF4472C4"
    assert ws["B1"].number_format == "@"
    assert ws["B2"].number_format == "0.00"
    assert ws["C2"].fill.fgColor.rgb == "FFDAE3F3"


@pytest.mark.parametrize("engine", ["xlsxwriter", "openpyxl"])
def test_save_formatted_xlsx_resolves_theme_color_references(workspace_tmp_dir: Path, engine: str) -> None:
    output = workspace_tmp_dir / f"theme_refs_{engine}.xlsx"
    data = [["Name", "Value"], ["Alpha", 1]]
    formatter = {
        "row": {
            0: {
                "pattern": "solid",
                "fg_color": "office_theme:Accent1+40",
                "font_color": "white",
                "bold": True,
            },
            -1: {},
        },
        "col": {},
    }

    saved_path = save_formatted_xlsx(
        data_table=data,
        file_name=output,
        sheet_name="Report",
        formatter=formatter,
        engine=engine,  # type: ignore[arg-type]
        mode="new",
        header=1,
    )

    wb = load_workbook(saved_path)
    ws = wb["Report"]

    assert ws["A1"].fill.fgColor.rgb == "FF8FAADC"
    assert ws["A1"].font.color is not None
    assert ws["A1"].font.color.rgb == "FFFFFFFF"


def test_replace_with_backup_retargets_supported_references(workspace_tmp_dir: Path) -> None:
    output = workspace_tmp_dir / "replace_refs.xlsx"

    from openpyxl import Workbook
    wb = Workbook()
    report = wb.active
    report.title = "Report"
    report["A1"] = 10
    calc = wb.create_sheet("Calc")
    calc["A1"] = "='Report (1)'!A1"

    dv = DataValidation(type="list", formula1="='Report (1)'!$A$1:$A$2")
    calc.add_data_validation(dv)
    dv.add("B1")

    wb.defined_names.add(DefinedName(name="ShiftedRef", attr_text="'Report (1)'!$A$1"))
    wb.save(output)

    save_formatted_xlsx(
        data_table=[["H"], [1]],
        file_name=output,
        sheet_name="Report",
        formatter={"row": {}, "col": {}},
        engine="openpyxl",
        mode="replace",
        backup=True,
        header=1,
    )

    wb2 = load_workbook(output)
    calc2 = wb2["Calc"]
    assert calc2["A1"].value == "=Report!A1"
    assert calc2.data_validations.dataValidation[0].formula1 == "=Report!$A$1:$A$2"

    shifted_ref = next(name for name in wb2.defined_names.values() if name.name == "ShiftedRef")
    assert shifted_ref.attr_text == "Report!$A$1"

