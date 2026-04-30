from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer


@dataclass
class RefEntry:
    kind: str
    location: str
    expression: str
    refs: list[str]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="inspect_sheet_references",
        description=(
            "Inspect workbook references that can be retargeted by sheetkit "
            "(cell formulas, defined names, data validations)."
        ),
    )
    parser.add_argument("file_name", help="Workbook .xlsx file to inspect.")
    parser.add_argument(
        "-sn",
        "--sheet-name",
        dest="sheet_name",
        default=None,
        help=(
            "Optional sheet name filter. If provided, output includes only entries that "
            "reference this sheet or its versioned variants like 'Sheet (1)'."
        ),
    )
    parser.add_argument(
        "-s",
        "--strict",
        action="store_true",
        help=(
            "Return non-zero exit code when matching references are found. "
            "With --sheet-name: checks only that sheet and its versioned names. "
            "Without --sheet-name: checks any supported cross-sheet references."
        ),
    )
    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Suppress console output (exit code still reflects result).",
    )
    parser.add_argument(
        "-log",
        "--log",
        dest="log_file",
        default=None,
        help="Append run protocol to a log file.",
    )
    return parser


def _emit(lines: list[str], *, quiet: bool, log_file: str | None) -> None:
    text = "\n".join(lines)
    if not quiet and text:
        print(text)
    if log_file:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{ts}] inspect_sheet_references\n")
            if text:
                f.write(text)
                f.write("\n")
            f.write("\n")


def _unquote_sheet_name(sheet: str) -> str:
    if len(sheet) >= 2 and sheet[0] == "'" and sheet[-1] == "'":
        return sheet[1:-1].replace("''", "'")
    return sheet


def _extract_sheet_refs(expr: str) -> list[str]:
    if not isinstance(expr, str) or "!" not in expr:
        return []

    formula = expr if expr.startswith("=") else f"={expr}"
    refs: list[str] = []

    try:
        tok = Tokenizer(formula)
    except Exception:
        return refs

    for item in tok.items:
        token_value = item.value
        if not isinstance(token_value, str) or "!" not in token_value:
            continue

        left = token_value.split("!", 1)[0]
        for part in left.split(":"):
            sheet_part = part.strip()
            if not sheet_part:
                continue

            if "]" in sheet_part and sheet_part.startswith("["):
                sheet_part = sheet_part[sheet_part.rfind("]") + 1:]

            sheet_name = _unquote_sheet_name(sheet_part)
            if sheet_name:
                refs.append(sheet_name)

    return refs


def _matches_sheet_filter(refs: list[str], sheet_name: str) -> bool:
    if not refs:
        return False
    version_prefix = f"{sheet_name} ("
    return any(ref == sheet_name or ref.startswith(version_prefix) for ref in refs)


def _collect_references(file_name: Path) -> list[RefEntry]:
    wb = load_workbook(file_name, data_only=False)
    entries: list[RefEntry] = []

    # 1) Cell formulas
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=") and "!" in value):
                    continue
                refs = _extract_sheet_refs(value)
                if refs:
                    entries.append(
                        RefEntry(
                            kind="formula",
                            location=f"{ws.title}!{cell.coordinate}",
                            expression=value,
                            refs=refs,
                        )
                    )

    # 2) Defined names
    for defined_name in wb.defined_names.values():
        attr_text = getattr(defined_name, "attr_text", None)
        if not (isinstance(attr_text, str) and "!" in attr_text):
            continue
        refs = _extract_sheet_refs(attr_text)
        if refs:
            entries.append(
                RefEntry(
                    kind="defined_name",
                    location=f"name:{defined_name.name}",
                    expression=attr_text,
                    refs=refs,
                )
            )

    # 3) Data validations
    for ws in wb.worksheets:
        data_validations = getattr(ws, "data_validations", None)
        validations = getattr(data_validations, "dataValidation", None)
        if validations is None:
            continue

        for idx, dv in enumerate(validations, start=1):
            sqref = str(getattr(dv, "sqref", "") or "")
            for attr in ("formula1", "formula2"):
                value = getattr(dv, attr, None)
                if not (isinstance(value, str) and "!" in value):
                    continue
                refs = _extract_sheet_refs(value)
                if refs:
                    entries.append(
                        RefEntry(
                            kind="data_validation",
                            location=f"{ws.title}!DV#{idx}:{attr}:{sqref}",
                            expression=value,
                            refs=refs,
                        )
                    )

    return entries


def main() -> int:
    parser = build_parser()
    raw_argv = ["--log" if token == "-log" else token for token in sys.argv[1:]]
    args = parser.parse_args(raw_argv)

    workbook_path = Path(args.file_name)
    if not workbook_path.exists():
        parser.error(f"Workbook not found: {workbook_path!s}")

    try:
        entries = _collect_references(workbook_path)
        if args.sheet_name:
            entries = [e for e in entries if _matches_sheet_filter(e.refs, args.sheet_name)]

        lines: list[str] = []
        if not entries:
            if args.sheet_name:
                lines.append(f"No supported references found for sheet: {args.sheet_name}")
            else:
                lines.append("No supported cross-sheet references found.")
            _emit(lines, quiet=args.quiet, log_file=args.log_file)
            return 0

        lines.append(f"Workbook: {workbook_path.resolve()}")
        if args.sheet_name:
            lines.append(f"Filter: sheet='{args.sheet_name}' (+ versioned names)")
        lines.append(f"Entries: {len(entries)}")
        lines.append("")

        target_hits = 0
        version_hits = 0
        for entry in entries:
            lines.append(f"[{entry.kind}] {entry.location}")
            lines.append(f"  refs: {', '.join(entry.refs)}")
            lines.append(f"  expr: {entry.expression}")
            if args.sheet_name:
                for ref in entry.refs:
                    if ref == args.sheet_name:
                        target_hits += 1
                    elif ref.startswith(f"{args.sheet_name} ("):
                        version_hits += 1
            lines.append("")

        if args.sheet_name:
            lines.append(f"Exact '{args.sheet_name}' refs: {target_hits}")
            lines.append(f"Versioned '{args.sheet_name} (N)' refs: {version_hits}")

        _emit(lines, quiet=args.quiet, log_file=args.log_file)

        if args.strict:
            return 1
        return 0
    except Exception as e:
        _emit([str(e)], quiet=args.quiet, log_file=args.log_file)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

