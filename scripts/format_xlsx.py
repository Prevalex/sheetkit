from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from sheetkit import load_formatter, write_sheet


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="format_xlsx",
        description="Format a csv/xls/xlsx file using an sheetkit formatter JSON.",
    )
    parser.add_argument("input_file", help="Input file: .csv, .xlsx, or .xls")
    parser.add_argument("output_file", help="Output .xlsx file")
    parser.add_argument(
        "-ft",
        "--formatter-file",
        dest="formatter_file",
        required=True,
        help="Formatter JSON file",
    )
    parser.add_argument(
        "-hr",
        "--header-rows",
        "--header",
        dest="header",
        type=int,
        default=None,
        choices=(0, 1, 2),
        help="Optional header row/column override: 0, 1, or 2. If omitted, formatter header styles are kept as-is.",
    )
    parser.add_argument(
        "-z",
        "--zebra",
        dest="zebra",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Optional zebra override. Use --zebra or --no-zebra.",
    )
    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Suppress console output.",
    )
    parser.add_argument(
        "-log",
        "--log",
        dest="log_file",
        default=None,
        help="Append run protocol to a log file.",
    )
    return parser


def _emit(lines: list[str], *, quiet: bool, log_file: str | None) -> None:
    text = "\n".join(lines)
    if not quiet and text:
        print(text)
    if log_file:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{ts}] format_xlsx\n")
            if text:
                f.write(text)
                f.write("\n")
            f.write("\n")


def _read_csv(path: Path) -> list[list[Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [list(row) for row in csv.reader(f)]


def _read_xlsx(path: Path) -> list[list[Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    data: list[list[Any]] = []
    max_col = ws.max_column
    for row in ws.iter_rows(min_col=1, max_col=max_col, values_only=True):
        data.append(list(row))
    while data and not data[-1]:
        data.pop()
    while data and all(value is None for value in data[-1]):
        data.pop()
    return data


def _read_xls(path: Path) -> list[list[Any]]:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as e:
        raise RuntimeError(
            "Reading .xls requires the optional dependency 'xlrd'. "
            "Install it in the current environment to use .xls input."
        ) from e

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    return [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]


def load_input_table(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    raise ValueError(f"Unsupported input format: {path.suffix!r}. Expected .csv, .xlsx, or .xls")


def main() -> int:
    parser = build_parser()
    raw_argv = ["--log" if token == "-log" else token for token in sys.argv[1:]]
    args = parser.parse_args(raw_argv)

    input_path = Path(args.input_file)
    output_path = Path(args.output_file)
    formatter_path = Path(args.formatter_file)

    try:

        data_table = load_input_table(input_path)
        formatter = load_formatter(formatter_path)
        priority = formatter.get("priority", "row")
        row_spec = dict(formatter.get("row", {}))
        col_spec = dict(formatter.get("col", {}))
        header_axis = col_spec if priority == "col" else row_spec
        if args.header is not None:
            if args.header < 2:
                header_axis.pop(1, None)
            if args.header < 1:
                header_axis.pop(0, None)
        if args.zebra is False:
            header_axis.pop(-2, None)
        formatter = {"priority": priority, "row": row_spec, "col": col_spec}

        save_kwargs = {
            "data_table": data_table,
            "file_name": output_path,
            "formatter": formatter,
            "engine": "xlsxwriter",
            "mode": "new",
        }
        if args.header is not None:
            save_kwargs["header"] = args.header

        result = write_sheet(**save_kwargs)
        _emit([str(result.resolve())], quiet=args.quiet, log_file=args.log_file)
        return 0

    except Exception as e:
        _emit([str(e)], quiet=args.quiet, log_file=args.log_file)
        return 1

if __name__ == "__main__":
    raise SystemExit(main())

