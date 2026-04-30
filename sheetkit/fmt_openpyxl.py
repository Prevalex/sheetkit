#!

import re

from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, Protection
from openpyxl.workbook.workbook import Workbook

from .colors import normalize_color_value
from .types import FormatDict
from .utils import map_style_value

#✅=============== [👉 ПОД_СЕКЦИЯ СПЕЦИАЛЬНОГО ФОРМАТИРОВАНИЯ OPENPYXL] ================

def _normalize_color_openpyxl(value):
    """
    Accepts a Boolean color value (HEX string, CSS name, etc.)
    and returns the string 'AARRGGBB' for openpyxl.
    """
    if value is None:
        return None
    # хотим 'AARRGGBB'
    return normalize_color_value(value, with_alpha=True)

class OpenpyxlCellStyle:
    __slots__ = ("alignment", "border", "fill", "font", "number_format", "protection")

    def __init__(self, alignment=None, border=None, fill=None,
                 font=None, number_format=None, protection=None):
        self.alignment = alignment
        self.border = border
        self.fill = fill
        self.font = font
        self.number_format = number_format
        self.protection = protection

def _make_openpyxl_style(
    fmt: FormatDict,
) -> OpenpyxlCellStyle:
    """
    Creates a set of openpyxl style objects (Alignment, Border, PatternFill, Font, Protection)
    based on a logical FormatDict.

    Color fields (font_color, fg_color, bg_color, border*_color, border_color)
    can be:
    - HEX: 'FFAACC', '#FFAACC', '80FFAACC'
    - CSS name: 'steelblue', 'red', ...
    """

    # ---------- Alignment ----------
    alignment = None
    if any(k in fmt for k in ("align", "valign", "text_wrap", "indent", "shrink_to_fit", "text_rotation")):
        align_kwargs = {}

        horiz = map_style_value("align", fmt.get("align"), "openpyxl")
        vert  = map_style_value("valign", fmt.get("valign"), "openpyxl")

        if horiz is not None:
            align_kwargs["horizontal"] = horiz
        if vert is not None:
            align_kwargs["vertical"] = vert

        if "text_wrap" in fmt and fmt["text_wrap"] is not None:
            align_kwargs["wrap_text"] = bool(fmt["text_wrap"])

        if "indent" in fmt and fmt["indent"] is not None:
            align_kwargs["indent"] = int(fmt["indent"])

        if "shrink_to_fit" in fmt and fmt["shrink_to_fit"] is not None:
            align_kwargs["shrink_to_fit"] = bool(fmt["shrink_to_fit"])

        if "text_rotation" in fmt and fmt["text_rotation"] is not None:
            align_kwargs["textRotation"] = int(fmt["text_rotation"])

        if align_kwargs:
            alignment = Alignment(**align_kwargs)

    # ---------- Border ----------
    border = None
    common_style  = fmt.get("border", None)
    common_color  = fmt.get("border_color", None)

    common_style_mapped = map_style_value("border", common_style, "openpyxl") if common_style is not None else None
    common_color_norm   = _normalize_color_openpyxl(common_color) if common_color is not None else None

    def side_style(name):
        val = fmt.get(name, None)
        if val is None:
            return None
        return map_style_value("border", val, "openpyxl")

    def side_color(name):
        col_val = fmt.get(name, None)
        if col_val is None:
            return common_color_norm
        return _normalize_color_openpyxl(col_val)

    left_style   = side_style("border_left")   or common_style_mapped
    right_style  = side_style("border_right")  or common_style_mapped
    top_style    = side_style("border_top")    or common_style_mapped
    bottom_style = side_style("border_bottom") or common_style_mapped

    left_color   = side_color("border_left_color")
    right_color  = side_color("border_right_color")
    top_color    = side_color("border_top_color")
    bottom_color = side_color("border_bottom_color")

    def make_side(style_value, color_value):
        if style_value is None and color_value is None:
            return None
        kw = {}
        if style_value is not None:
            kw["style"] = style_value
        if color_value is not None:
            kw["color"] = color_value
        if not kw:
            return None
        return Side(**kw)

    left_side   = make_side(left_style,   left_color)
    right_side  = make_side(right_style,  right_color)
    top_side    = make_side(top_style,    top_color)
    bottom_side = make_side(bottom_style, bottom_color)

    if any((left_side, right_side, top_side, bottom_side)):
        border = Border(
            left=left_side   or Side(),
            right=right_side or Side(),
            top=top_side     or Side(),
            bottom=bottom_side or Side(),
        )

    # ---------- Fill ----------
    fill = None
    if "pattern" in fmt or "fg_color" in fmt or "bg_color" in fmt:
        pattern_type = map_style_value("pattern", fmt.get("pattern"), "openpyxl")
        fg_color_raw = fmt.get("fg_color", None)
        bg_color_raw = fmt.get("bg_color", None)

        fg_color = _normalize_color_openpyxl(fg_color_raw) if fg_color_raw is not None else None
        bg_color = _normalize_color_openpyxl(bg_color_raw) if bg_color_raw is not None else None

        fill_kwargs = {}
        if pattern_type is not None:
            fill_kwargs["patternType"] = pattern_type
        elif fg_color is not None or bg_color is not None:
            fill_kwargs["patternType"] = "solid"

        if fg_color is not None:
            fill_kwargs["fgColor"] = fg_color
        if bg_color is not None:
            fill_kwargs["bgColor"] = bg_color

        if fill_kwargs:
            fill = PatternFill(**fill_kwargs)

    # ---------- Font ----------
    font = None
    if any(k in fmt for k in ("bold", "italic", "underline", "strike", "font_color", "font_name", "font_size")):
        font_kwargs = {}

        if "bold" in fmt and fmt["bold"] is not None:
            font_kwargs["bold"] = bool(fmt["bold"])
        if "italic" in fmt and fmt["italic"] is not None:
            font_kwargs["italic"] = bool(fmt["italic"])
        if "strike" in fmt and fmt["strike"] is not None:
            font_kwargs["strike"] = bool(fmt["strike"])

        if "underline" in fmt and fmt["underline"] is not None:
            val = fmt["underline"]
            mapped = map_style_value("underline", val, "openpyxl")
            font_kwargs["underline"] = mapped

        if "font_color" in fmt and fmt["font_color"] is not None:
            col = _normalize_color_openpyxl(fmt["font_color"])
            if col is not None:
                font_kwargs["color"] = col

        if "font_name" in fmt and fmt["font_name"] is not None:
            font_kwargs["name"] = fmt["font_name"]
        if "font_size" in fmt and fmt["font_size"] is not None:
            font_kwargs["size"] = fmt["font_size"]

        if font_kwargs:
            font = Font(**font_kwargs)

    # ---------- Number format ----------
    number_format = fmt.get("num_format", None)

    # ---------- Protection ----------
    protection = None
    if "locked" in fmt or "hidden" in fmt:
        prot_kwargs = {}
        if "locked" in fmt and fmt["locked"] is not None:
            prot_kwargs["locked"] = bool(fmt["locked"])
        if "hidden" in fmt and fmt["hidden"] is not None:
            prot_kwargs["hidden"] = bool(fmt["hidden"])
        if prot_kwargs:
            protection = Protection(**prot_kwargs)

    return OpenpyxlCellStyle(
        alignment=alignment,
        border=border,
        fill=fill,
        font=font,
        number_format=number_format,
        protection=protection,
    )


def apply_style(cell, style: OpenpyxlCellStyle):
    if style.alignment is not None:
        cell.alignment = style.alignment
    if style.border is not None:
        cell.border = style.border
    if style.fill is not None:
        cell.fill = style.fill
    if style.font is not None:
        cell.font = style.font
    if style.number_format is not None:
        cell.number_format = style.number_format
    if style.protection is not None:
        cell.protection = style.protection

################ [ XLSX VIEW CONTROL SECTION ] ##############
# Sheet order, sheet renaming, column widths, etc. ##

# Sheet versioning helper
def rotate_sheet_versions(wb: Workbook, base_name: str, copy_sheet=False) -> str | None:
    """
    Renames the existing sheet base_name to base_name(N),
    where N = 1 + the maximum number of existing versions.

    If the sheet base_name does not exist, it does nothing.
    That is: This function:
    returns nothing (simply renames);
    safely creates the "next version" of the sheet, without breaking previous versions.

    * If copy_sheet=True, base_name is not renamed. Instead, a copy is created
    with the name f"{base_name}({max_i + 1})", and the sheet itself remains. This is necessary for cases where
    this sheet is later edited in the program, rather than created.

    Example:
    the sheets are: ['Report', 'Report (1)', 'Report (3)']
    rotate_sheet_versions(wb, 'Report')
    -> 'Report' is renamed to 'Report (4)'
    """
    if base_name not in wb.sheetnames:
        return None

    existing_names = wb.sheetnames
    pattern = re.compile(rf"^{re.escape(base_name)} \((\d+)\)$")
    max_i = 0
    for name in existing_names:
        m = pattern.match(name)
        if m:
            idx = int(m.group(1))
            if idx > max_i:
                max_i = idx
    else:
        ws = wb[base_name]
        sheet_index = wb.worksheets.index(ws) + 1
        if copy_sheet:
            # Copy the worksheet
            ws = wb.copy_worksheet(ws)
            offset = sheet_index - wb.worksheets.index(ws)
            if offset != 0:
                wb.move_sheet(ws, offset)

        new_name = f"{base_name} ({max_i + 1})"
        ws.title = new_name
        return new_name


def _quote_sheet_name(name: str) -> str:
    """
    Excel-style quoting for worksheet names.
    """
    needs_quotes = any(ch in " [](){}+-*/&^%'!?:;" for ch in name)
    escaped = name.replace("'", "''")
    if needs_quotes:
        return f"'{escaped}'"
    return escaped


def _unquote_sheet_name(s: str) -> str:
    """
    Removes Excel single-quote wrapping from a sheet name.
    """
    if len(s) >= 2 and s[0] == "'" and s[-1] == "'":
        inner = s[1:-1]
        return inner.replace("''", "'")
    return s


def _replace_sheet_in_token_value(
        token_value: str,
        old_name: str,
        new_name: str,
) -> str:
    """
    Replaces a worksheet name inside a single formula token, preserving syntax.
    """
    if "!" not in token_value:
        return token_value

    left, right = token_value.split("!", 1)
    parts = left.split(":")
    new_parts: list[str] = []

    for part in parts:
        part = part.strip()
        book_prefix = ""
        sheet_part = part
        if "]" in part and part.startswith("["):
            idx = part.rfind("]")
            book_prefix = part[: idx + 1]
            sheet_part = part[idx + 1:]

        sheet_unquoted = _unquote_sheet_name(sheet_part)
        if sheet_unquoted == old_name:
            if len(sheet_part) >= 2 and sheet_part[0] == "'" and sheet_part[-1] == "'":
                quoted = _quote_sheet_name(new_name)
                new_part = f"{book_prefix}{quoted}"
            else:
                maybe_quoted = _quote_sheet_name(new_name)
                new_part = f"{book_prefix}{maybe_quoted}"
        else:
            new_part = part
        new_parts.append(new_part)

    new_left = ":".join(new_parts)
    return f"{new_left}!{right}"


def _retarget_formula_text(
        formula: str,
        old_name: str,
        new_name: str,
) -> str:
    """
    Retargets sheet references old_name -> new_name inside an Excel formula string.
    """
    from openpyxl.formula.tokenizer import Tokenizer

    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula

    tok = Tokenizer(formula)
    changed = False
    for t in tok.items:
        if isinstance(t.value, str) and "!" in t.value:
            new_val = _replace_sheet_in_token_value(t.value, old_name, new_name)
            if new_val != t.value:
                t.value = new_val
                changed = True
    if not changed:
        return formula
    rebuilt = "".join(t.value for t in tok.items)
    if formula.startswith("=") and not rebuilt.startswith("="):
        rebuilt = f"={rebuilt}"
    return rebuilt


def retarget_sheet_references(
        wb: Workbook,
        old_name: str,
        new_name: str,
        *,
        exclude_sheets: set[str] | None = None,
) -> None:
    """
    Retargets references old_name -> new_name in workbook objects that are supported by openpyxl:
    - cell formulas
    - defined names
    - data validations (formula1/formula2)
    """
    if not old_name or not new_name or old_name == new_name:
        return
    excluded = exclude_sheets or set()

    # 1) Cell formulas
    for sheet in wb.worksheets:
        if sheet.title in excluded:
            continue
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and "!" in value:
                    cell.value = _retarget_formula_text(value, old_name, new_name)

    # 2) Defined names
    for defined_name in wb.defined_names.values():
        attr_text = getattr(defined_name, "attr_text", None)
        if not isinstance(attr_text, str) or "!" not in attr_text:
            continue
        original_has_eq = attr_text.startswith("=")
        expr = attr_text if original_has_eq else f"={attr_text}"
        patched = _retarget_formula_text(expr, old_name, new_name)
        if patched != expr:
            defined_name.attr_text = patched if original_has_eq else patched[1:]

    # 3) Data validations
    for sheet in wb.worksheets:
        if sheet.title in excluded:
            continue
        data_validations = getattr(sheet, "data_validations", None)
        validation_items = getattr(data_validations, "dataValidation", None)
        if validation_items is None:
            continue
        for dv in validation_items:
            for attr in ("formula1", "formula2"):
                val = getattr(dv, attr, None)
                if not isinstance(val, str) or "!" not in val:
                    continue
                has_eq = val.startswith("=")
                expr = val if has_eq else f"={val}"
                patched = _retarget_formula_text(expr, old_name, new_name)
                if patched != expr:
                    setattr(dv, attr, patched if has_eq else patched[1:])


# Safely rename a sheet in an xlsx file (openpyxl)
def rename_sheet_safely(
    wb: Workbook,
    old_name: str,
    new_name: str,
    update_formulas: bool = True,
) -> None:
    """
    Rename the sheet old_name -> new_name and, if necessary, update the formulas in the workbook so that references to the
    sheet don't turn into #REF.

    Requirement: OpenPyXL doesn't adjust sheet references when renaming a sheet. Only Excel does this.
    Therefore, you must go through all sheets and manually change the sheet name if it appears
    in formula references.

    Supports:
    - Regular references: Sheet!A1
    - Quotation marks: 'My Sheet'!A1
    - External workbooks: [Book.xlsx]Sheet!A1
    - Partial 3D references: Sheet1:Sheet3!A1

    Limitations:
    - Does not edit:
    * Formulas in Defined Names
    * Formulas in charts, conditional formatting, and data validations
    (these can be added separately, if needed).
    """

    if old_name not in wb.sheetnames:
        raise KeyError(f"Sheet {old_name!r} not found in workbook")

    if old_name == new_name:
        return

    # 1) Rename the sheet itself
    ws = wb[old_name]
    ws.title = new_name

    if not update_formulas:
        return

    #2) Update formulas in all sheets
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue

                new_formula = _retarget_formula_text(v, old_name, new_name)
                if new_formula != v:
                    cell.value = new_formula
