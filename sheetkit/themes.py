#!

import json
import os
from pathlib import Path
from os import PathLike
from zipfile import ZipFile
from collections.abc import Mapping
from typing import Any, Literal
import re
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils.cell import coordinate_to_tuple

from .types import AxisFormatSpec, FormatDict, FormatPriorityLiteral, SheetFormatSpec
from .utils import _compact_fmt, _guess_align_from_num_format
from .colors import _mix_hex_colors, _excel_accent_percent, _pick_contrast_color

#
# ############### [ RUNTIME CACHE OF IMPORTED THEMES] ###########
#

# Theme presets live in presets/themes/*.json.
# THEMES is kept as an in-memory cache for themes imported during runtime.
THEMES: dict[str, dict[str, Any]] = {}

USER_THEME_NAME_PREFIX = "_"
DEFAULT_OFFICE_THEME_DIR = Path(r"C:\Program Files\Microsoft Office\root\Document Themes 16")
DEFAULT_USER_THEME_DIR = Path(os.environ.get("APPDATA", "")) / "Microsoft" / "Templates" / "Document Themes"
PRESETS_DIR = Path(__file__).resolve().parent / "presets"
THEMES_DIR = PRESETS_DIR / "themes"
FORMATTERS_DIR = PRESETS_DIR / "formatters"
THEME_KIND = "themes"
FORMATTER_KIND = "formatter"
THEME_XML_PATH = "theme/theme/theme1.xml"
FORMATTER_COLOR_KEYS = {
    "fg_color",
    "bg_color",
    "font_color",
    "border_left_color",
    "border_right_color",
    "border_top_color",
    "border_bottom_color",
}
COLOR_NAME_MAP = {
    "dk1": "dark1",
    "lt1": "light1",
    "dk2": "dark2",
    "lt2": "light2",
    "accent1": "accent1",
    "accent2": "accent2",
    "accent3": "accent3",
    "accent4": "accent4",
    "accent5": "accent5",
    "accent6": "accent6",
    "hlink": "hlink",
    "folHlink": "folHlink",
}


#
# ############### [ HELPERS] ##############
#

def _normalize_theme_key(value: str) -> str:
    """
    Casts an arbitrary theme name to the THEMES dictionary key.
    """
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def _normalize_preset_key(file_name: str) -> str:
    """
    Converts the name of a preset file/resource to a canonical key.
    """
    return _normalize_theme_key(Path(str(file_name)).stem)


def _load_json_file(file_name: str | PathLike) -> dict[str, Any]:
    """
    Loads a JSON file and returns a dictionary.
    """
    path = Path(file_name)
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"Preset file {path!s} must contain a JSON object")
    return data


def _find_preset_file(directory: Path, name: str) -> Path | None:
    """
    Searches for a preset file by name in the specified directory.
    """
    preset_key = _normalize_preset_key(name)
    for json_file in directory.glob("*.json"):
        if _normalize_preset_key(json_file.stem) == preset_key:
            return json_file
    return None


def _get_child_local_name(elem: ET.Element) -> str:
    """
    Returns the local-name of an XML tag without the namespace.
    """
    return elem.tag.rsplit("}", 1)[-1]


def _extract_theme_color_value(color_holder: ET.Element) -> str | None:
    """
    Extracts the color value from the dk1/lt1/accentN/hlink/folHlink node.

    Priority:
    - srgbClr[@val]
    - sysClr[@lastClr]
    - sysClr[@val] as a fallback
    """
    for child in color_holder:
        local_name = _get_child_local_name(child)
        if local_name == "srgbClr":
            value = child.attrib.get("val")
            if value:
                return value.upper()
        elif local_name == "sysClr":
            value = child.attrib.get("lastClr") or child.attrib.get("val")
            if value:
                return value.upper()
    return None


def _extract_theme_color_details(color_holder: ET.Element) -> tuple[str | None, dict[str, Any] | None]:
    """
    Extracts both resolved color value and transform metadata from a color node.
    """
    for child in color_holder:
        color_type = _get_child_local_name(child)
        if color_type not in ("srgbClr", "sysClr"):
            continue

        if color_type == "srgbClr":
            resolved_value = child.attrib.get("val")
        else:
            resolved_value = child.attrib.get("lastClr") or child.attrib.get("val")

        if not resolved_value:
            continue

        transforms: list[dict[str, Any]] = []
        for transform in child:
            op_name = _get_child_local_name(transform)
            op_payload: dict[str, Any] = {"op": op_name}
            for k, v in transform.attrib.items():
                op_payload[k] = v
            transforms.append(op_payload)

        details: dict[str, Any] = {
            "color_type": color_type,
            "resolved_value": resolved_value.upper(),
            "transforms": transforms,
        }

        if color_type == "srgbClr":
            details["value"] = child.attrib.get("val")
        else:
            details["value"] = child.attrib.get("val")
            details["lastClr"] = child.attrib.get("lastClr")

        return resolved_value.upper(), details

    return None, None


def _parse_theme_xml(theme_xml: str) -> dict[str, Any]:
    """
    Parses the DrawingML theme from theme1.xml and extracts colors/fonts.
    """
    root = ET.fromstring(theme_xml)

    theme_name = root.attrib.get("name", "")
    colors: dict[str, str] = {}
    color_transforms: dict[str, dict[str, Any]] = {}
    fonts: dict[str, str] = {}

    clr_scheme = root.find(".//{*}clrScheme")
    if clr_scheme is not None:
        for child in clr_scheme:
            source_name = _get_child_local_name(child)
            target_name = COLOR_NAME_MAP.get(source_name)
            if not target_name:
                continue
            value, details = _extract_theme_color_details(child)
            if value:
                colors[target_name] = value
            if details is not None:
                color_transforms[target_name] = details

    font_scheme = root.find(".//{*}fontScheme")
    if font_scheme is not None:
        major_font = font_scheme.find("./{*}majorFont/{*}latin")
        minor_font = font_scheme.find("./{*}minorFont/{*}latin")
        if major_font is not None and major_font.attrib.get("typeface"):
            fonts["major"] = major_font.attrib["typeface"]
        if minor_font is not None and minor_font.attrib.get("typeface"):
            fonts["minor"] = minor_font.attrib["typeface"]

    return {
        "scheme_name": theme_name,
        "colors": colors,
        "color_transforms": color_transforms,
        "fonts": fonts,
    }


def _read_zip_text(zip_path: str | PathLike, entry_name: str) -> str:
    """
    Reads a text file from an Office theme zip package.
    """
    with ZipFile(zip_path) as archive:
        with archive.open(entry_name) as file_obj:
            return file_obj.read().decode("utf-8")


def _extract_thmx_app_info(zip_path: str | PathLike) -> dict[str, str]:
    """
    Tries to extract application metadata from .thmx docProps/app.xml.
    """
    info: dict[str, str] = {}
    try:
        app_xml = _read_zip_text(zip_path, "docProps/app.xml")
    except KeyError:
        return info

    try:
        root = ET.fromstring(app_xml)
    except ET.ParseError:
        return info

    application = root.find(".//{*}Application")
    if application is not None and application.text is not None:
        app_name = application.text.strip()
        if app_name:
            info["application"] = app_name

    app_version = root.find(".//{*}AppVersion")
    if app_version is not None and app_version.text is not None:
        version = app_version.text.strip()
        if version:
            info["version"] = version

    return info


def _iter_theme_variant_paths(zip_path: str | PathLike) -> list[str]:
    """
    Returns paths to variant*/theme/theme/theme1.xml within .thmx.
    """
    with ZipFile(zip_path) as archive:
        return sorted(
            entry.filename
            for entry in archive.infolist()
            if entry.filename.startswith("themeVariants/variant")
            and entry.filename.endswith("/theme/theme/theme1.xml")
        )


def _get_theme_by_name(name: str) -> dict[str, Any]:
    """
    Returns a built-in theme preset by name from presets/themes.
    """
    preset_file = _find_preset_file(THEMES_DIR, name)
    if preset_file is None:
        raise ValueError(f"Theme preset {name!r} not found")
    return load_preset_file(preset_file, expected_kind=THEME_KIND)


def _compose_formatter(
        *,
        row: Mapping[int, FormatDict],
        col: Mapping[int, FormatDict],
        priority: FormatPriorityLiteral,
) -> SheetFormatSpec:
    """
    Builds a canonical runtime SheetFormatSpec dictionary.
    """
    return {
        "priority": priority,
        "row": dict(row),
        "col": dict(col),
    }


def _resolve_workbook_sheet(wb: Any, sheet_name: str | None) -> Any:
    """
    Returns a workbook sheet, either by name or wb.active.
    """
    if sheet_name is None:
        return wb.active
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet {sheet_name!r} not found")
    return wb[sheet_name]


def _normalize_imported_hex_color(color: str | None) -> str | None:
    """
    Normalizes color from openpyxl to RRGGBB format.
    """
    if not color:
        return None
    value = str(color).strip().lstrip("#").upper()
    if len(value) == 8:
        if value == "00000000":
            return None
        return value[2:]
    if len(value) == 6:
        return value
    return None


def _get_workbook_theme_colors(wb: Any) -> list[str | None]:
    """
    Returns a list of the workbook theme's base colors in openpyxl index order.
    """
    loaded_theme = getattr(wb, "loaded_theme", None)
    if not loaded_theme:
        return []

    if isinstance(loaded_theme, bytes):
        theme_xml = loaded_theme.decode("utf-8")
    else:
        theme_xml = str(loaded_theme)

    theme_info = _parse_theme_xml(theme_xml)
    colors = theme_info.get("colors", {})
    if not isinstance(colors, Mapping):
        return []

    # openpyxl theme indexes: bg1, tx1, bg2, tx2, accent1..6, hlink, folHlink
    return [
        colors.get("light1"),
        colors.get("dark1"),
        colors.get("light2"),
        colors.get("dark2"),
        colors.get("accent1"),
        colors.get("accent2"),
        colors.get("accent3"),
        colors.get("accent4"),
        colors.get("accent5"),
        colors.get("accent6"),
        colors.get("hlink"),
        colors.get("folHlink"),
    ]


def _apply_openpyxl_tint(base_color: str, tint: float) -> str:
    """
    Applies a simple approximation of Excel tint to the base color.
    """
    if tint == 0:
        return base_color
    if tint > 0:
        return _mix_hex_colors(base_color, "FFFFFF", min(tint, 1.0))
    return _mix_hex_colors(base_color, "000000", min(abs(tint), 1.0))


def _resolve_openpyxl_color(color: Any, theme_colors: list[str | None]) -> str | None:
    """
    Resolves openpyxl.styles.colors.Color to RRGGBB.
    """
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return _normalize_imported_hex_color(getattr(color, "rgb", None))
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if isinstance(indexed, int) and 0 <= indexed < len(COLOR_INDEX):
            return _normalize_imported_hex_color(COLOR_INDEX[indexed])
        return None
    if color_type == "theme":
        theme_idx = getattr(color, "theme", None)
        if isinstance(theme_idx, int) and 0 <= theme_idx < len(theme_colors):
            base_color = theme_colors[theme_idx]
            if base_color:
                tint = float(getattr(color, "tint", 0.0) or 0.0)
                return _apply_openpyxl_tint(base_color, tint)
        return None
    if color_type == "auto":
        return None
    return _normalize_imported_hex_color(getattr(color, "value", None))


def _map_openpyxl_border_style(style: str | None) -> int | None:
    """
    Reduces openpyxl border styles to a simplified sheetkit logical model.
    """
    if not style:
        return None
    if style == "medium":
        return 2
    return 1


def _cell_fill_to_format(cell: Any, theme_colors: list[str | None]) -> FormatDict:
    """
    Extracts the cell's fill parameters into our FormatDict.
    """
    fill = cell.fill
    if fill is None:
        return {}

    pattern_type = getattr(fill, "patternType", None) or getattr(fill, "fill_type", None)
    fg_color = _resolve_openpyxl_color(getattr(fill, "fgColor", None), theme_colors)
    bg_color = _resolve_openpyxl_color(getattr(fill, "bgColor", None), theme_colors)
    if pattern_type == "solid" and bg_color == "000000":
        bg_color = None

    if pattern_type != "solid" and not fg_color and not bg_color:
        return {}

    fmt: FormatDict = {}
    if pattern_type == "solid" or fg_color:
        fmt["pattern"] = "solid"
    if fg_color:
        fmt["fg_color"] = fg_color
    if bg_color:
        fmt["bg_color"] = bg_color
    elif fmt.get("pattern") == "solid" and fg_color:
        # Для solid fill в preset лучше сохранять и bg_color тоже.
        # Это делает стиль самодостаточным и не дает row[-1] "подтекать"
        # через наследование в zebra row (-2), что особенно заметно в xlsxwriter.
        fmt["bg_color"] = fg_color
    return fmt


def _cell_font_to_format(cell: Any, theme_colors: list[str | None]) -> FormatDict:
    """
    Retrieves the font properties of a cell.
    """
    font = cell.font
    if font is None:
        return {}

    underline = getattr(font, "underline", None)
    if underline == "none":
        underline = None

    return _compact_fmt(
        {
            "font_name": getattr(font, "name", None),
            "font_size": getattr(font, "size", None),
            "bold": getattr(font, "bold", None),
            "italic": getattr(font, "italic", None),
            "underline": underline,
            "strike": getattr(font, "strike", None),
            "font_color": _resolve_openpyxl_color(getattr(font, "color", None), theme_colors),
        }
    )


def _cell_alignment_to_format(cell: Any) -> FormatDict:
    """
    Retrieves the alignment properties of a cell.
    """
    alignment = cell.alignment
    if alignment is None:
        return {}

    valign = getattr(alignment, "vertical", None)
    if valign == "centerContinuous":
        valign = "center"

    return _compact_fmt(
        {
            "align": getattr(alignment, "horizontal", None),
            "valign": valign,
            "text_wrap": getattr(alignment, "wrap_text", None),
        }
    )


def _cell_border_to_format(cell: Any, theme_colors: list[str | None]) -> FormatDict:
    """
    Retrieves the border properties of a cell.
    """
    border = cell.border
    if border is None:
        return {}

    fmt: FormatDict = {}
    for side_name, fmt_key in (
            ("left", "border_left"),
            ("right", "border_right"),
            ("top", "border_top"),
            ("bottom", "border_bottom"),
    ):
        side = getattr(border, side_name, None)
        style_value = _map_openpyxl_border_style(getattr(side, "style", None))
        color_value = _resolve_openpyxl_color(getattr(side, "color", None), theme_colors)
        if style_value is not None:
            fmt[fmt_key] = style_value
        if color_value is not None:
            fmt[f"{fmt_key}_color"] = color_value
    return fmt


def _cell_to_row_style(cell: Any, theme_colors: list[str | None]) -> FormatDict:
    """
    Assembles a simplified logical row format from a single reference cell.
    """
    return _compact_fmt(
        {
            **_cell_fill_to_format(cell, theme_colors),
            **_cell_font_to_format(cell, theme_colors),
            **_cell_alignment_to_format(cell),
            **_cell_border_to_format(cell, theme_colors),
        }
    )


def _cell_num_format(cell: Any) -> str | None:
    """
    Returns the cell's num_format if it is meaningful.
    """
    num_format = getattr(cell, "number_format", None)
    if not num_format or num_format == "General":
        return None
    return str(num_format)


def _extract_column_num_formats(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        columns: int,
        header_rows: int,
        zebra: bool,
) -> dict[int, FormatDict]:
    """
    Extracts num_format by columns. Priority: data rows, then header rows.
    """
    by_index: dict[int, FormatDict] = {}
    data_row_numbers = [start_row + header_rows]
    if zebra:
        data_row_numbers.append(start_row + header_rows + 1)
    header_row_numbers = [start_row + offset for offset in range(header_rows)]
    probe_rows = data_row_numbers + header_row_numbers

    for col_offset in range(columns):
        col_num_format: str | None = None
        for row_number in probe_rows:
            num_format = _cell_num_format(ws.cell(row=row_number, column=start_col + col_offset))
            if num_format is not None:
                col_num_format = num_format
                break
        if col_num_format is not None:
            by_index[col_offset] = {"num_format": col_num_format}

    return by_index


def _extract_row_num_formats(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        rows: int,
        header_rows: int,
        zebra: bool,
) -> dict[int, FormatDict]:
    """
    Extracts num_format by rows. Priority: data columns, then header columns.
    """
    by_index: dict[int, FormatDict] = {}
    data_col_numbers = [start_col + header_rows]
    if zebra:
        data_col_numbers.append(start_col + header_rows + 1)
    header_col_numbers = [start_col + offset for offset in range(header_rows)]
    probe_cols = data_col_numbers + header_col_numbers

    for row_offset in range(rows):
        row_num_format: str | None = None
        for col_number in probe_cols:
            num_format = _cell_num_format(ws.cell(row=start_row + row_offset, column=col_number))
            if num_format is not None:
                row_num_format = num_format
                break
        if row_num_format is not None:
            by_index[row_offset] = {"num_format": row_num_format}

    return by_index


def _extract_column_num_formats_from_range(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        rows: int,
        columns: int,
) -> dict[int, FormatDict]:
    """
    Extracts column num_format values by scanning each column down a rectangular range.
    """
    by_index: dict[int, FormatDict] = {}
    for col_offset in range(columns):
        col_num_format: str | None = None
        for row_offset in range(rows):
            num_format = _cell_num_format(
                ws.cell(row=start_row + row_offset, column=start_col + col_offset)
            )
            if num_format is not None:
                col_num_format = num_format
                break
        if col_num_format is not None:
            by_index[col_offset] = {"num_format": col_num_format}
    return by_index


def _extract_row_num_formats_from_range(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        rows: int,
        columns: int,
) -> dict[int, FormatDict]:
    """
    Extracts row num_format values by scanning each row across a rectangular range.
    """
    by_index: dict[int, FormatDict] = {}
    for row_offset in range(rows):
        row_num_format: str | None = None
        for col_offset in range(columns):
            num_format = _cell_num_format(
                ws.cell(row=start_row + row_offset, column=start_col + col_offset)
            )
            if num_format is not None:
                row_num_format = num_format
                break
        if row_num_format is not None:
            by_index[row_offset] = {"num_format": row_num_format}
    return by_index


def _build_formatter_from_sheet_samples(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        columns: int | None,
        rows: int | None,
        header_rows: int,
        zebra: bool,
        priority: FormatPriorityLiteral,
        theme_colors: list[str | None],
) -> SheetFormatSpec:
    """
    Assembles a runtime formatter from a marked template range.
    """
    row_spec: AxisFormatSpec = {}
    col_spec: AxisFormatSpec = {}

    if priority == "col":
        if rows is None:
            raise ValueError("rows must be specified when priority='col'")
        if header_rows >= 1:
            col_spec[0] = _cell_to_row_style(ws.cell(row=start_row, column=start_col), theme_colors)
        if header_rows >= 2:
            col_spec[1] = _cell_to_row_style(ws.cell(row=start_row, column=start_col + 1), theme_colors)

        cols_base_index = start_col + header_rows
        col_spec[-1] = _cell_to_row_style(ws.cell(row=start_row, column=cols_base_index), theme_colors)
        if zebra:
            col_spec[-2] = _cell_to_row_style(ws.cell(row=start_row, column=cols_base_index + 1), theme_colors)

        row_spec.update(
            _extract_row_num_formats(
                ws,
                start_row=start_row,
                start_col=start_col,
                rows=rows,
                header_rows=header_rows,
                zebra=zebra,
            )
        )
        return _compose_formatter(row=row_spec, col=col_spec, priority=priority)

    if columns is None:
        raise ValueError("columns must be specified when priority='row'")
    if header_rows >= 1:
        row_spec[0] = _cell_to_row_style(ws.cell(row=start_row, column=start_col), theme_colors)
    if header_rows >= 2:
        row_spec[1] = _cell_to_row_style(ws.cell(row=start_row + 1, column=start_col), theme_colors)

    rows_base_index = start_row + header_rows
    row_spec[-1] = _cell_to_row_style(ws.cell(row=rows_base_index, column=start_col), theme_colors)
    if zebra:
        row_spec[-2] = _cell_to_row_style(ws.cell(row=rows_base_index + 1, column=start_col), theme_colors)

    col_spec.update(
        _extract_column_num_formats(
            ws,
            start_row=start_row,
            start_col=start_col,
            columns=columns,
            header_rows=header_rows,
            zebra=zebra,
        )
    )
    return _compose_formatter(row=row_spec, col=col_spec, priority=priority)


def _build_formatter_from_sheet_range(
        ws: Any,
        *,
        start_row: int,
        start_col: int,
        rows: int,
        columns: int,
        priority: FormatPriorityLiteral,
        theme_colors: list[str | None],
) -> SheetFormatSpec:
    """
    Assembles a formatter by copying row or column styles from a rectangular range.
    """
    row_spec: AxisFormatSpec = {}
    col_spec: AxisFormatSpec = {}

    if priority == "col":
        for col_offset in range(columns):
            fmt = _cell_to_row_style(ws.cell(row=start_row, column=start_col + col_offset), theme_colors)
            target_idx = -1 if col_offset == columns - 1 else col_offset
            col_spec[target_idx] = fmt
        row_spec.update(
            _extract_row_num_formats_from_range(
                ws,
                start_row=start_row,
                start_col=start_col,
                rows=rows,
                columns=columns,
            )
        )
        return _compose_formatter(row=row_spec, col=col_spec, priority=priority)

    for row_offset in range(rows):
        fmt = _cell_to_row_style(ws.cell(row=start_row + row_offset, column=start_col), theme_colors)
        target_idx = -1 if row_offset == rows - 1 else row_offset
        row_spec[target_idx] = fmt
    col_spec.update(
        _extract_column_num_formats_from_range(
            ws,
            start_row=start_row,
            start_col=start_col,
            rows=rows,
            columns=columns,
        )
    )
    return _compose_formatter(row=row_spec, col=col_spec, priority=priority)


def _find_excel_theme_file(theme: str, theme_dir: str | PathLike | None = None) -> Path | None:
    """
    Searches for a .thmx file by theme name.

    If theme_dir is None, searches default locations in this order:
    1) DEFAULT_USER_THEME_DIR
    2) DEFAULT_OFFICE_THEME_DIR
    """
    theme_key = _normalize_theme_key(theme)

    if theme_dir is None:
        search_dirs: list[Path] = [DEFAULT_USER_THEME_DIR, DEFAULT_OFFICE_THEME_DIR]
    else:
        search_dirs = [Path(theme_dir)]

    for theme_dir_path in search_dirs:
        if not theme_dir_path.exists():
            continue
        for thmx_file in theme_dir_path.glob("*.thmx"):
            if _normalize_theme_key(thmx_file.stem) == theme_key:
                return thmx_file

    return None


def _get_theme_font_defaults(
        theme_info: Mapping[str, Any],
        *,
        font_name: str | None,
        font_size: int | None,
) -> tuple[str, int]:
    """
    Resolves default font name and size for a theme.
    """
    fonts = theme_info.get("fonts", {})

    resolved_font_name = font_name or fonts.get("minor") or fonts.get("major") or "Calibri"
    resolved_font_size = 11 if font_size is None else font_size

    return resolved_font_name, resolved_font_size


def _resolve_theme_colors(
        theme: Mapping[str, Any],
        *,
        variant: int = 0,
) -> Mapping[str, Any]:
    """
    Resolves colors mapping from theme with optional variant selection.

    Rules:
    - variant=0: base colors
    - variant>=1: uses theme["variants"][variant-1] if available
    - missing/invalid variants: silently fall back to base colors
    """
    base_colors_obj = theme.get("colors", theme)
    base_colors = base_colors_obj if isinstance(base_colors_obj, Mapping) else {}

    if variant < 1:
        return base_colors

    variants_obj = theme.get("variants")
    if not isinstance(variants_obj, list):
        return base_colors

    variant_idx = variant - 1
    if variant_idx < 0 or variant_idx >= len(variants_obj):
        return base_colors

    variant_item = variants_obj[variant_idx]
    if not isinstance(variant_item, Mapping):
        return base_colors

    variant_colors_obj = variant_item.get("colors")
    if not isinstance(variant_colors_obj, Mapping):
        return base_colors

    merged_colors: dict[str, Any] = dict(base_colors)
    merged_colors.update(variant_colors_obj)
    return merged_colors


def _theme_ref(
        theme: Mapping[str, Any],
        ref: str,
        *,
        color_mode: Literal["hex", "ref"],
) -> str:
    """
    Returns a theme color reference, prefixed with theme name when available.
    """
    if color_mode != "ref":
        return ref

    theme_name = theme.get("name")
    if isinstance(theme_name, str) and theme_name:
        return f"{theme_name}:{ref}"
    return ref


def _apply_types(
        base_spec: SheetFormatSpec,
        *,
        types: list[str | None] | None,
        priority: FormatPriorityLiteral,
) -> SheetFormatSpec:
    """
    Adds num_format values from types to a SheetFormatSpec.
    """
    axis_key = "row" if priority == "col" else "col"
    axis_spec: AxisFormatSpec = dict(base_spec.get(axis_key, {}))

    if types:
        for idx, num_fmt in enumerate(types):
            if num_fmt is None:
                continue

            fmt: FormatDict = {"num_format": num_fmt}

            # Alignment heuristics by num_format
            align = _guess_align_from_num_format(num_fmt)
            fmt["align"] = align

            # It is possible to add something else in the future, for example shrink_to_fit for long texts.
            if idx in axis_spec:
                axis_spec[idx] = {**axis_spec[idx], **fmt}
            else:
                axis_spec[idx] = fmt

    base_spec[axis_key] = axis_spec

    return base_spec


def iter_thmx_files(theme_dir: str | PathLike) -> list[Path]:
    """
    Returns a list of .thmx files from a directory.
    """
    theme_dir_path = Path(theme_dir)
    if not theme_dir_path.exists():
        return []
    return sorted(theme_dir_path.glob("*.thmx"))


def load_thmx_theme(file_name: str | PathLike) -> dict[str, Any]:
    """
    Loads a theme from a .thmx file.

    Returns a dictionary in a THEMES-compatible format:
    {
     "kind": "themes",
     "source": "Theme.thmx",
     "application": "...",  # optional
     "version": "...",  # optional
     "scheme_name": "...",
     "colors": {...},
     "fonts": {"major": "...", "minor": "..."},
     "variants": [...]
    }
    """
    path = Path(file_name)
    if not path.exists():
        raise FileNotFoundError(path)

    parsed_theme_info = _parse_theme_xml(_read_zip_text(path, THEME_XML_PATH))
    theme_info: dict[str, Any] = {
        "kind": THEME_KIND,
        "source": path.name,
    }
    theme_info.update(_extract_thmx_app_info(path))
    theme_info.update(parsed_theme_info)

    variants: list[dict[str, Any]] = []
    for variant_path in _iter_theme_variant_paths(path):
        variant_info = _parse_theme_xml(_read_zip_text(path, variant_path))
        if variant_info.get("colors"):
            variants.append(variant_info)

    if variants:
        theme_info["variants"] = variants

    return theme_info


def make_theme_preset_name(theme_name: str, *, is_user_theme: bool = False) -> str:
    """
    Returns the canonical name of the theme's preset file.
    For custom themes, the _ prefix is added if necessary.
    """
    base_name = _normalize_theme_key(theme_name)
    if is_user_theme:
        return f"{USER_THEME_NAME_PREFIX + base_name}"
    return base_name


def get_theme(theme: str | PathLike, *, auto_import: bool = True) -> dict[str, Any]:
    """
    Returns a theme by name or explicit path.

    Resolution for theme name (str):
    1) THEMES_DIR / f"{theme}.json"
    2) DEFAULT_USER_THEME_DIR / f"{theme}.thmx" (if folder exists)
    3) DEFAULT_OFFICE_THEME_DIR / f"{theme}.thmx" (if folder exists)

    Explicit path supports:
    - .json: loaded via load_preset_file(...)
    - .thmx: loaded via import_theme(...)
    """
    theme_str = str(theme)
    theme_path = Path(theme_str)

    # Explicit path mode (PathLike or path-like str with directory part)
    is_explicit_path = (
            (not isinstance(theme, str))
            or theme_path.is_absolute()
            or theme_path.parent != Path(".")
    )
    if is_explicit_path:
        if not theme_path.exists() or not theme_path.is_file():
            raise ValueError(f"Theme file {theme_path!s} not found")
        suffix = theme_path.suffix.lower()
        if suffix == ".json":
            return load_preset_file(theme_path, expected_kind=THEME_KIND)
        if suffix == ".thmx":
            return import_theme(theme_path)
        raise ValueError(f"Unsupported theme file extension: {theme_path.suffix!r}")

    # Name mode: prefer built-in json preset first
    preset_file = THEMES_DIR / f"{theme_str}.json"
    if preset_file.exists():
        return load_preset_file(preset_file, expected_kind=THEME_KIND)
    try:
        return _get_theme_by_name(theme_str)
    except ValueError:
        pass

    if auto_import:
        found_file = _find_excel_theme_file(theme_str, theme_dir=DEFAULT_USER_THEME_DIR)
        if found_file is not None:
            return import_theme(found_file)

        found_file = _find_excel_theme_file(theme_str, theme_dir=DEFAULT_OFFICE_THEME_DIR)
        if found_file is not None:
            return import_theme(found_file)

    # Runtime cache fallback
    normalized_key = _normalize_theme_key(theme_str)
    if theme_str in THEMES:
        return THEMES[theme_str]
    if normalized_key in THEMES:
        return THEMES[normalized_key]

    raise ValueError(f"Theme {theme!r} not found in presets/user/office/cache")


def register_theme(
        theme_info: Mapping[str, Any],
        *,
        theme_name: str | None = None,
        replace: bool = True,
) -> str:
    """
    Registers the imported theme in the THEMES dictionary and returns its key.
    """
    if theme_name is None:
        if "scheme_name" not in theme_info:
            raise ValueError("theme_info must contain 'scheme_name' when theme_name is not provided")
        theme_name = str(theme_info["scheme_name"])

    key = _normalize_theme_key(theme_name)

    if not replace and key in THEMES:
        return key

    THEMES[key] = dict(theme_info)
    return key


def import_theme(
        file_name: str | PathLike,
        *,
        theme_name: str | None = None,
        register: bool = True,
        replace: bool = True,
) -> dict[str, Any]:
    """
    Imports a single .thmx theme.

    `file_name` may be:
    - direct path to .thmx file
    - theme name (searched first in DEFAULT_USER_THEME_DIR, then in DEFAULT_OFFICE_THEME_DIR)

    If register=True, the theme is also added to the THEMES directory.
    """
    input_path = Path(file_name)
    resolved_file: Path | None = None

    if input_path.exists() and input_path.is_file():
        resolved_file = input_path
    else:
        lookup_name = input_path.stem if input_path.suffix.lower() == ".thmx" else str(file_name)
        resolved_file = _find_excel_theme_file(lookup_name)

    if resolved_file is None:
        raise FileNotFoundError(input_path)

    theme_info = load_thmx_theme(resolved_file)
    if register:
        key = register_theme(theme_info, theme_name=theme_name, replace=replace)
        return THEMES[key]
    return theme_info


def import_themes(
        theme_dir: str | PathLike | None = None,
        *,
        replace: bool = True,
) -> dict[str, dict[str, Any]]:
    """
    Imports .thmx files and registers them in THEMES.

    If theme_dir is None, searches default locations in this order:
    1) DEFAULT_OFFICE_THEME_DIR
    2) DEFAULT_USER_THEME_DIR

    This keeps user themes as final winners on key collisions.
    """
    imported: dict[str, dict[str, Any]] = {}
    if theme_dir is None:
        search_dirs: list[Path] = [DEFAULT_OFFICE_THEME_DIR, DEFAULT_USER_THEME_DIR]
    else:
        search_dirs = [Path(theme_dir)]

    if theme_dir is not None and not search_dirs[0].exists():
        raise FileNotFoundError(search_dirs[0])

    for theme_dir_path in search_dirs:
        if not theme_dir_path.exists():
            continue

        for thmx_file in sorted(theme_dir_path.glob("*.thmx")):
            imported_theme = import_theme(thmx_file, replace=replace)
            theme_key = _normalize_theme_key(imported_theme.get("scheme_name", thmx_file.stem))
            imported[theme_key] = imported_theme

    return imported


def build_formatter_from_theme(
        theme: Mapping[str, Any],
        *,
        header: int | None = None,
        zebra: bool = True,
        variant: int = 0,
        accent: int = 1,
        priority: FormatPriorityLiteral = "row",
        color_mode: Literal["hex", "ref"] = "hex",
        types: list[str | None] | None = None,
        font_name: str | None = None,
        font_size: int | None = None,
        border_style: int | None = None,
) -> SheetFormatSpec:
    """
    Builds a SheetFormatSpec based on theme colors. It accepts one theme (full object or colors dictionary) and
     builds SheetFormatSpec.

    The theme parameter can be:

    - a full theme object from presets/themes JSON:
        {
        "file_name": "...",
        "scheme_name": "Facet",
        "colors": { "dark1": "...", ... }
        }

    - or simply a dictionary of colors:
        { "dark1": "000000", "light1": "FFFFFF", ... }

    In the resulting SheetFormatSpec:
    - axis[0] — the style of the first header row/column (if header >= 1)
    - axis[1] — the style of the second header row/column (if header >= 2)
    - axis[-1] — the default row/column style
    - axis[-2] — the second zebra row/column style (if zebra=True)

    Color scheme:
    - header1:
      bg = selected accent, font_color = light1, border_bottom = dark2
    - header2:
      bg = selected accent 60% (Excel-like), font_color = light1
    - baselines:
      -1: transparent/no fill
      -2: bg = selected accent 20% (Excel-like)
    """
    if priority not in ("row", "col"):
        raise ValueError("priority must be 'row' or 'col'")
    if color_mode not in {"hex", "ref"}:
        raise ValueError("color_mode must be 'hex' or 'ref'")
    header_count = 1 if header is None else header
    if header_count < 0:
        raise ValueError("header must be >= 0")

    # 1) Get a color dictionary (optionally from a specific theme variant)
    colors = _resolve_theme_colors(theme, variant=variant)

    dark1 = colors.get("dark1", "000000")
    light1 = colors.get("light1", "FFFFFF")
    dark2 = colors.get("dark2", dark1)
    accent_key = f"accent{accent}" if isinstance(accent, int) and 1 <= accent <= 6 else "accent1"
    accent_value = colors.get(accent_key)
    if not isinstance(accent_value, str) or not accent_value:
        accent_value = colors.get("accent1", dark2)
        accent_key = "accent1"
    accent_base = str(accent_value)

    # 2) Basic font/border defaults for generated styles
    resolved_font_name, resolved_font_size = _get_theme_font_defaults(
        theme,
        font_name=font_name,
        font_size=font_size,
    )
    resolved_border_style = 1 if border_style is None else border_style

    # 2) Basic font style for all lines
    base_font_fmt: FormatDict = _compact_fmt(
        {
            "font_name": resolved_font_name,
            "font_size": resolved_font_size,
        }
    )

    # 3) Zebra Rows
    # - Row -1: Transparent/no fill
    # - Row -2: Accent1 20% (Excel-like)
    zebra_row_even = _compact_fmt(
        {
            **base_font_fmt,
            "border_bottom": resolved_border_style,
            "border_bottom_color": _theme_ref(theme, "Dark2", color_mode=color_mode) if color_mode == "ref" else dark2,
        }
    )

    # In theme-ref mode we use +80 to match Excel-like "20% accent" appearance.
    zebra_alt_bg = (
        _theme_ref(theme, f"{accent_key.title()}+80", color_mode=color_mode)
        if color_mode == "ref"
        else _excel_accent_percent(accent_base, 20)
    )
    zebra_row_odd = _compact_fmt(
        {
            **base_font_fmt,
            "fg_color": zebra_alt_bg,
            "border_bottom": resolved_border_style,
            "border_bottom_color": _theme_ref(theme, "Dark2", color_mode=color_mode) if color_mode == "ref" else dark2,
        }
    )

    row_spec: AxisFormatSpec = {}
    col_spec: AxisFormatSpec = {}
    primary_spec = col_spec if priority == "col" else row_spec

    # -1: default (e.g. even rows/columns)
    primary_spec[-1] = zebra_row_even

    # -2: Second zebra row/column (if enabled)
    if zebra:
        primary_spec[-2] = zebra_row_odd

    # 4) Headers
    header1_fmt = _compact_fmt(
        {
            **base_font_fmt,
            "bold": True,
            "align": "center",
            "valign": "center",
            "text_wrap": True,
            "fg_color": _theme_ref(theme, accent_key.title(), color_mode=color_mode) if color_mode == "ref" else accent_base,
            "font_color": _theme_ref(theme, "Light1", color_mode=color_mode) if color_mode == "ref" else light1,
            "border_bottom": resolved_border_style,
            "border_bottom_color": _theme_ref(theme, "Dark2", color_mode=color_mode) if color_mode == "ref" else dark2,
        }
    )

    # Second header line is Accent1 60% (Excel-like).
    # In theme-ref mode we use +40 to match Excel-like "60% accent" appearance.
    header2_bg = (
        _theme_ref(theme, f"{accent_key.title()}+40", color_mode=color_mode)
        if color_mode == "ref"
        else _excel_accent_percent(accent_base, 60)
    )
    header2_bg_hex = _excel_accent_percent(accent_base, 60)
    header2_font_color = _pick_contrast_color(header2_bg_hex, light_color=light1, dark_color=dark1)
    header2_font_color_out = "Light1" if header2_font_color.upper() == str(light1).upper() else "Dark1"
    if color_mode == "ref":
        header2_font_color_out = _theme_ref(theme, header2_font_color_out, color_mode=color_mode)

    header2_fmt = _compact_fmt(
        {
            **base_font_fmt,
            "bold": True,
            "align": "center",
            "valign": "center",
            "text_wrap": True,
            "fg_color": header2_bg,
            "font_color": header2_font_color_out if color_mode == "ref" else header2_font_color,
            "border_bottom": resolved_border_style,
            "border_bottom_color": _theme_ref(theme, "Dark2", color_mode=color_mode) if color_mode == "ref" else dark2,
        }
    )

    # header:
    #  0 — no explicit header styles
    #  1 — first row/column
    #  2 — first two rows/columns
    if header_count >= 1:
        primary_spec[0] = header1_fmt

    if header_count >= 2:
        primary_spec[1] = header2_fmt

    sheet_spec: SheetFormatSpec = {
        "priority": priority,
        "col": col_spec,
        "row": row_spec,
    }

    return _apply_types(sheet_spec, types=types, priority=priority)


def load_preset_file(
        source: str | PathLike,
        *,
        expected_kind: str | None = None,
) -> dict[str, Any]:
    """
    Loads a preset JSON file.
    """
    preset = _load_json_file(source)
    preset.setdefault("name", _normalize_preset_key(str(source)))
    if expected_kind is not None:
        kind_value = preset.get("kind")
        if kind_value is None:
            raise ValueError(f"Preset file {source!s} must contain key 'kind'={expected_kind!r}")
        if kind_value != expected_kind:
            raise ValueError(
                f"Preset file {source!s} has invalid kind {kind_value!r}; expected {expected_kind!r}"
            )
    return preset


def save_preset_file(
        preset: Mapping[str, Any],
        json_file: str | PathLike,
) -> Path:
    """
    Saves a preset dictionary to a JSON file.
    """
    json_path = Path(json_file)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(dict(preset), f, ensure_ascii=False, indent=2)
        f.write("\n")
    return json_path


def _to_axis_format_spec(axis_data: Any, *, axis_name: str) -> AxisFormatSpec:
    """
    Converts an arbitrary mapping with string/int keys to AxisFormatSpec.
    """
    if not isinstance(axis_data, Mapping):
        raise ValueError(f"Formatter field '{axis_name}' must be a mapping")

    axis_spec: AxisFormatSpec = {}
    for raw_idx, raw_fmt in axis_data.items():
        try:
            idx = int(raw_idx)
        except (TypeError, ValueError) as e:
            raise ValueError(f"Formatter field '{axis_name}' contains non-integer key: {raw_idx!r}") from e
        if not isinstance(raw_fmt, Mapping):
            raise ValueError(f"Formatter field '{axis_name}[{raw_idx}]' must be a mapping")
        axis_spec[idx] = dict(raw_fmt)
    return axis_spec


def _normalize_formatter_priority(value: Any) -> FormatPriorityLiteral:
    """
    Normalizes formatter priority, preserving row-priority for existing files.
    """
    return "col" if value == "col" else "row"


def list_preset_themes() -> list[str]:
    """
    Returns sorted preset theme names from presets/themes.
    """
    names = {
        _normalize_preset_key(json_file.stem)
        for json_file in THEMES_DIR.glob("*.json")
    }
    return sorted(name for name in names if name)


def list_preset_formatters() -> list[str]:
    """
    Returns sorted preset formatter names that have both _row and _col variants.
    """
    row_names: set[str] = set()
    col_names: set[str] = set()

    for json_file in FORMATTERS_DIR.glob("*.json"):
        stem = _normalize_preset_key(json_file.stem)
        if stem.endswith("_row"):
            row_names.add(stem[:-4])
        elif stem.endswith("_col"):
            col_names.add(stem[:-4])

    return sorted(name for name in (row_names & col_names) if name)


def _resolve_header_count(*, header: int | None, header_rows: int | None, default: int) -> int:
    """
    Resolves the new header name and the older header_rows alias.
    """
    if header is not None and header_rows is not None and header != header_rows:
        raise ValueError("header and header_rows must not specify different values")
    value = header if header is not None else header_rows
    return default if value is None else value


def load_formatter(source: str | PathLike) -> SheetFormatSpec:
    """
    Loads a formatter JSON file to runtime SheetFormatSpec.
    """
    payload = load_preset_file(source)
    kind_value = payload.get("kind")
    if kind_value is not None and kind_value != FORMATTER_KIND:
        raise ValueError(f"Formatter file {source!s} has invalid kind {kind_value!r}; expected {FORMATTER_KIND!r}")
    if "row" not in payload or "col" not in payload:
        raise ValueError("Formatter JSON must contain 'row' and 'col' fields")

    embedded_theme = payload.get("theme")
    if isinstance(embedded_theme, Mapping):
        theme_name = embedded_theme.get("name") or payload.get("name")
        register_theme(embedded_theme, theme_name=str(theme_name) if theme_name else None, replace=True)

    return {
        "priority": _normalize_formatter_priority(payload.get("priority")),
        "row": _to_axis_format_spec(payload["row"], axis_name="row"),
        "col": _to_axis_format_spec(payload["col"], axis_name="col"),
    }


def load_format_preset(
        name: str,
        priority: FormatPriorityLiteral = "row",
) -> SheetFormatSpec:
    """
    Loads a formatter preset by name from presets/formatters only.
    """
    normalized_priority = _normalize_formatter_priority(priority)
    preset_file = _find_preset_file(FORMATTERS_DIR, f"{name}_{normalized_priority}")
    if preset_file is None:
        raise ValueError(
            f"Formatter preset {name!r} with priority {normalized_priority!r} not found in {FORMATTERS_DIR!s}"
        )
    return load_formatter(preset_file)


def save_formatter(
        formatter: Mapping[str, Any],
        json_file: str | PathLike,
        *,
        name: str | None = None,
        theme: Mapping[str, Any] | None = None,
) -> Path:
    """
    Saves a runtime formatter (`SheetFormatSpec`) to JSON.
    """
    normalized_formatter = {
        "row": _to_axis_format_spec(formatter.get("row", {}), axis_name="row"),
        "col": _to_axis_format_spec(formatter.get("col", {}), axis_name="col"),
    }
    payload: dict[str, Any] = {
        "kind": FORMATTER_KIND,
        "priority": _normalize_formatter_priority(formatter.get("priority")),
        "row": normalized_formatter["row"],
        "col": normalized_formatter["col"],
    }
    if name:
        payload["name"] = name
    if theme is not None:
        payload["theme"] = dict(theme)
    return save_preset_file(payload, json_file)


def resolve_formatter_colors(
        formatter: Mapping[str, Any],
        *,
        theme: str | Mapping[str, Any] | None = None,
) -> SheetFormatSpec:
    """
    Resolves color references in formatter to HEX strings.
    """
    from .colors import normalize_color_value

    row_spec = _to_axis_format_spec(formatter.get("row", {}), axis_name="row")
    col_spec = _to_axis_format_spec(formatter.get("col", {}), axis_name="col")

    def _resolve_fmt(fmt: Mapping[str, Any]) -> FormatDict:
        resolved: FormatDict = {}
        for key, value in fmt.items():
            if key in FORMATTER_COLOR_KEYS and value is not None:
                resolved[key] = normalize_color_value(value, theme=theme)
            else:
                resolved[key] = value
        return resolved

    return {
        "priority": _normalize_formatter_priority(formatter.get("priority")),
        "row": {idx: _resolve_fmt(fmt) for idx, fmt in row_spec.items()},
        "col": {idx: _resolve_fmt(fmt) for idx, fmt in col_spec.items()},
    }


def extract_formatter_from_sheet(
        file_name: str | PathLike,
        sheet_name: str | None = None,
        *,
        columns: int | None = None,
        rows: int | None = None,
        start_cell: str = "A1",
        header: int | None = None,
        header_rows: int | None = None,
        zebra: bool = True,
        priority: FormatPriorityLiteral = "row",
) -> SheetFormatSpec:
    """
    Extracts a runtime formatter from a manually formatted worksheet.
    """
    if priority not in ("row", "col"):
        raise ValueError("priority must be 'row' or 'col'")
    if priority == "row" and (columns is None or columns < 1):
        raise ValueError("columns must be >= 1 when priority='row'")
    if priority == "col" and (rows is None or rows < 1):
        raise ValueError("rows must be >= 1 when priority='col'")
    header_count = _resolve_header_count(header=header, header_rows=header_rows, default=2)
    if header_count < 0:
        raise ValueError("header must be >= 0")

    start_row, start_col = coordinate_to_tuple(start_cell)

    wb = load_workbook(filename=file_name)
    ws = _resolve_workbook_sheet(wb, sheet_name)

    if priority == "col":
        required_last_col = start_col + header_count + (2 if zebra else 1) - 1
        if ws.max_column < required_last_col:
            raise ValueError("Worksheet does not contain enough sample columns for formatter import")
    else:
        required_last_row = start_row + header_count + (2 if zebra else 1) - 1
        if ws.max_row < required_last_row:
            raise ValueError("Worksheet does not contain enough sample rows for formatter import")

    theme_colors = _get_workbook_theme_colors(wb)
    return _build_formatter_from_sheet_samples(
        ws,
        start_row=start_row,
        start_col=start_col,
        columns=columns,
        rows=rows,
        header_rows=header_count,
        zebra=zebra,
        priority=priority,
        theme_colors=theme_colors,
    )


def extract_formatter_range_from_sheet(
        file_name: str | PathLike,
        sheet_name: str | None = None,
        *,
        columns: int,
        rows: int,
        start_cell: str = "A1",
        priority: FormatPriorityLiteral = "row",
) -> SheetFormatSpec:
    """
    Extracts a formatter from a fully styled rectangular worksheet range.

    Unlike extract_formatter_from_sheet(), this function does not interpret the
    range as header/base/zebra samples. It copies the style of each row or column:

    - priority="row": row styles are read from the first column; num_format is read by columns.
    - priority="col": column styles are read from the first row; num_format is read by rows.

    The last row/column style becomes the default style (-1).
    """
    if priority not in ("row", "col"):
        raise ValueError("priority must be 'row' or 'col'")
    if rows < 1:
        raise ValueError("rows must be >= 1")
    if columns < 1:
        raise ValueError("columns must be >= 1")

    start_row, start_col = coordinate_to_tuple(start_cell)

    wb = load_workbook(filename=file_name)
    ws = _resolve_workbook_sheet(wb, sheet_name)

    required_last_row = start_row + rows - 1
    required_last_col = start_col + columns - 1
    if ws.max_row < required_last_row:
        raise ValueError("Worksheet does not contain enough rows for formatter range extraction")
    if ws.max_column < required_last_col:
        raise ValueError("Worksheet does not contain enough columns for formatter range extraction")

    theme_colors = _get_workbook_theme_colors(wb)
    return _build_formatter_from_sheet_range(
        ws,
        start_row=start_row,
        start_col=start_col,
        rows=rows,
        columns=columns,
        priority=priority,
        theme_colors=theme_colors,
    )


def extract_formatter_range_to_file(
        file_name: str | PathLike,
        json_file: str | PathLike,
        sheet_name: str | None = None,
        *,
        columns: int,
        rows: int,
        start_cell: str = "A1",
        priority: FormatPriorityLiteral = "row",
        name: str | None = None,
) -> Path:
    """
    Extracts a full-range formatter and saves it as formatter JSON.
    """
    formatter = extract_formatter_range_from_sheet(
        file_name,
        sheet_name,
        columns=columns,
        rows=rows,
        start_cell=start_cell,
        priority=priority,
    )
    return save_formatter(formatter, json_file, name=name)


def extract_formatter_to_file(
        file_name: str | PathLike,
        json_file: str | PathLike,
        sheet_name: str | None = None,
        *,
        columns: int | None = None,
        rows: int | None = None,
        start_cell: str = "A1",
        header: int | None = None,
        header_rows: int | None = None,
        zebra: bool = True,
        priority: FormatPriorityLiteral = "row",
        name: str | None = None,
) -> Path:
    """
    Extracts a formatter from a worksheet and saves it as JSON.
    """
    formatter = extract_formatter_from_sheet(
        file_name,
        sheet_name,
        columns=columns,
        rows=rows,
        start_cell=start_cell,
        header=header,
        header_rows=header_rows,
        zebra=zebra,
        priority=priority,
    )
    return save_formatter(formatter, json_file, name=name)

