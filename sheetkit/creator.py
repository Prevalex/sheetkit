#!

from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from os import PathLike
from typing import Any, cast

from .const import MAX_COLUMN_WIDTH, TRY_OPTIMIZE, COLUMN_WIDTH_RESERVE_CHARS
from .formatter import build_row_formats_with_columns, translate_formatter
from .utils import compute_column_widths, clear_range_openpyxl
from .fmt_openpyxl import rotate_sheet_versions, retarget_sheet_references, apply_style
from .helpers import PercentProgress, wrn, add_ext, inspect_name
from .types import EngineLiteral, FormatPriorityLiteral, ModeLiteral, ResolvedEngineLiteral, RowFormats, SheetFormatSpec

SheetCoord = tuple[int, int]
SheetRangeTuple = tuple[SheetCoord, SheetCoord]
SheetRangeInput = str | SheetRangeTuple
PatchRangeInput = SheetRangeInput

#
################ [ СЕКЦИЯ СОХРАНЕНИЯ И ФОРМАТИРОВАНИЯ ЛИСТА EXCEL ]##############
#

def _infer_num_format_from_value(value: Any) -> str:
    """
    Infers num_format for a single cell value.
    """
    if isinstance(value, bool):
        return "General"
    if isinstance(value, int):
        return "#,##0"
    if isinstance(value, float):
        return "#,##0.00"
    if isinstance(value, (datetime, date)):
        return "dd.mm.yyyy"
    return "General"


def _ensure_runtime_inferred_num_formats(
        effective_formatter: SheetFormatSpec,
        *,
        data_table: list[list[Any]],
        header_rows: int,
        priority: FormatPriorityLiteral,
        num_rows: int,
        num_cols: int,
) -> None:
    """
    Adds inferred num_format values at runtime when they are missing.
    """
    if num_cols <= 0 or num_rows <= 0:
        return

    if priority == "col":
        sample_col_idx = header_rows
        if sample_col_idx < 0 or sample_col_idx >= num_cols:
            return

        row_spec = _normalize_axis_spec(effective_formatter.setdefault("row", {}))
        default_num_format = row_spec.get(-1, {}).get("num_format")
        if isinstance(default_num_format, str) and default_num_format:
            effective_formatter["row"] = row_spec
            return

        for row_idx in range(num_rows):
            row_fmt = row_spec.get(row_idx, {})
            existing_num_format = row_fmt.get("num_format")
            if isinstance(existing_num_format, str) and existing_num_format:
                continue

            row = data_table[row_idx] if row_idx < len(data_table) else []
            value = row[sample_col_idx] if sample_col_idx < len(row) else None
            updated_row_fmt = dict(row_fmt)
            updated_row_fmt["num_format"] = _infer_num_format_from_value(value)
            row_spec[row_idx] = updated_row_fmt

        effective_formatter["row"] = row_spec
        return

    sample_row_idx = header_rows
    if sample_row_idx < 0 or sample_row_idx >= len(data_table):
        return

    sample_row = data_table[sample_row_idx]
    detached_col_spec = _normalize_axis_spec(effective_formatter.setdefault("col", {}))

    default_num_format = detached_col_spec.get(-1, {}).get("num_format")
    if isinstance(default_num_format, str) and default_num_format:
        effective_formatter["col"] = detached_col_spec
        return

    for col_idx in range(num_cols):
        col_fmt = detached_col_spec.get(col_idx, {})
        existing_num_format = col_fmt.get("num_format")
        if isinstance(existing_num_format, str) and existing_num_format:
            continue

        value = sample_row[col_idx] if col_idx < len(sample_row) else None
        inferred_num_format = _infer_num_format_from_value(value)

        updated_col_fmt = dict(col_fmt)
        updated_col_fmt["num_format"] = inferred_num_format
        detached_col_spec[col_idx] = updated_col_fmt

    effective_formatter["col"] = detached_col_spec


def _normalize_axis_spec(axis_spec: Any) -> dict[int, dict[str, Any]]:
    """
    Normalizes JSON-loaded axis keys to integers.
    """
    if not isinstance(axis_spec, dict):
        return {}

    normalized: dict[int, dict[str, Any]] = {}
    for raw_key, raw_fmt in axis_spec.items():
        if not isinstance(raw_fmt, dict):
            continue
        if isinstance(raw_key, int):
            key = raw_key
        elif isinstance(raw_key, str) and raw_key.lstrip("-").isdigit():
            key = int(raw_key)
        else:
            continue
        normalized[key] = dict(raw_fmt)
    return normalized


def _normalize_priority(value: Any) -> FormatPriorityLiteral:
    """
    Returns formatter priority, defaulting to row-priority for existing formatters.
    """
    return "col" if value == "col" else "row"


def _normalize_runtime_header_formats(
        axis_spec: dict[int, dict[str, Any]],
        *,
        header_count: int,
        axis_size: int,
        default_header_style: dict[str, Any],
) -> None:
    """
    Makes the save-time header argument authoritative for the priority axis.

    Header styles live in indexes 0 and 1 by convention. At save time, header_count
    decides how many of them are active; missing styles are copied from the last
    available header style, or from the default header style if no header style exists.
    """
    if axis_size <= 0:
        return

    existing_header_styles = [
        deepcopy(axis_spec[idx])
        for idx in range(min(2, axis_size))
        if idx in axis_spec
    ]

    if header_count <= 0:
        for idx in range(min(2, axis_size)):
            axis_spec.pop(idx, None)
        return

    active_count = min(header_count, axis_size)
    if existing_header_styles:
        header_templates = existing_header_styles
    else:
        header_templates = [deepcopy(default_header_style)]

    for idx in range(active_count):
        template_idx = min(idx, len(header_templates) - 1)
        axis_spec[idx] = deepcopy(header_templates[template_idx])

    for idx in range(active_count, min(2, axis_size)):
        axis_spec.pop(idx, None)


def write_sheet(
        data_table: list[list[Any]],
        file_name: str | PathLike,
        formatter: SheetFormatSpec | None = None,
        *,
        header: int | None = None,
        sheet_name: str | None = None,
        mode: ModeLiteral = "auto",
        engine: EngineLiteral = "auto",
        backup: bool = True,
        patch_skip_none: bool = True,
        patch_skip_formulas: bool = True,
        patch_skip_locked: bool = True,
        patch_range: PatchRangeInput | None = None,
        patch_strict_range: bool = False,
        as_template: bool = False,
        progressor: PercentProgress | None = None,
) -> Path:
    """
    Saves a table (list of lists) to an xlsx file with formatting.

    Args:
        data_table: Data table (list of rows, each row is a list of values).
        file_name: Full filename.
        sheet_name: Sheet name or None.
        formatter: Format specification (SheetFormatSpec) or None.
            Important:
            - row[-1] — Default row format.
            - row[-2] — Second default row format for "zebra".
            - row[i>=0] — Explicit row formats.

        engine: 'auto', 'openpyxl', or 'xlsxwriter'.
        mode:
            'new' — Create a new file. Raises FileExistsError if the file already exists.
            'replace' - ('openpyxl' only):
                * if backup=True:
                - rename the current sheet sheet_name to sheet_name(N)
                - create a new sheet sheet_name (new version).
                * if backup=False:
                - delete sheet sheet_name and create a new sheet sheet_name.

            'update' - ('openpyxl' only)
            * if backup=True:
            - create a copy of the specified sheet sheet_name to sheet_name(N)
            - overwrite the specified formats and data in the specified sheet
            * if backup=False:
            - overwrite the specified formats and data in the specified sheet

            'patch' - ('openpyxl' only)
            - updates only addressed cells from data_table in the existing target sheet.
            - by default:
                * None values are skipped (cell left unchanged),
                * formula cells are skipped,
                * locked cells are skipped when the sheet is protected.
            - existing styles, formulas, validations and layout are preserved.

            'auto' - if the file does not exist, the 'new' mode is selected;
                if the file exists, the 'replace' mode is selected.
                With engine='auto', xlsxwriter is used only for new workbooks;
                existing workbooks, replace/update/patch modes, and templates use openpyxl.

        header:
            0 — no header;
            1 — first row/column is considered a header;
            2 — first two rows/columns are considered a header.

            If header > 0, then a header format (bold, center, wrap) is logically
            created for indexes 0 and/or 1 on the priority axis,
            but only if the corresponding row index
            is not already explicitly specified in formatter.

        backup:
            Used in 'replace' mode for openpyxl.
            True — save the sheet as version sheet_name (N) before replacing it.
            False — do not create a backup (the old sheet is deleted).

        patch_skip_none:
            For mode='patch': if True, None values in data_table do not overwrite existing cells.
        patch_skip_formulas:
            For mode='patch': if True, cells containing formulas are not overwritten.
        patch_skip_locked:
            For mode='patch': if True, locked cells are not overwritten on protected sheets.
        patch_range:
            For mode='patch': optional range as:
            - Excel string (for example, 'D2:E7'), or
            - zero-based tuple coordinates:
              ((start_row, start_col), (end_row, end_col)).
            If set, data_table is anchored at the top-left corner of this range.
        patch_strict_range:
            For mode='patch': if True and patch_range is set, raises ValueError
            when data_table does not fit fully into the range.

        as_template: If True, the file is saved in template format with the .xltx extension.

        progressor: An object to which the file's row number is passed during processing. Access is performed
                    using the format progressor.update(row_index). Used, for example, to display progress/percentage
                    completion for large files.

    Returns:
        Path - the path to the saved file.

    Raises:
        ValueError, TypeError, OSError, and openpyxl/xlsxwriter library exceptions
        for invalid arguments or saving errors.
    """

    requested_mode = mode
    normalized_requested_mode = requested_mode.lower()
    if normalized_requested_mode == "auto" and patch_range is not None:
        # If caller asks for patch_range and target workbook already exists,
        # prefer patch semantics over replace semantics in auto mode.
        candidate_path = Path(add_ext(file_name, ".xltx" if as_template else ".xlsx"))
        if candidate_path.exists():
            requested_mode = "patch"

    engine, eff_mode, path = _resolve_engine_mode_and_path(
        file_name=file_name,
        engine=engine,
        mode=requested_mode,
        as_template=as_template,
    )

    num_rows = len(data_table)
    num_cols = max((len(row) for row in data_table), default=0)
    if header is None:
        header_count = 0
        apply_runtime_header_control = False
    else:
        if header < 0:
            raise ValueError("header must be >= 0")
        header_count = header
        apply_runtime_header_control = True

    # --------- Preparing an extended logical formatter (header lines) ---------
    if formatter is None:
        effective_formatter: SheetFormatSpec = {"col": {}, "row": {}}
    else:
        effective_formatter = {
            "col": dict(formatter.get("col", {})),
            "row": dict(formatter.get("row", {})),
            "priority": _normalize_priority(formatter.get("priority")),
        }
    priority = _normalize_priority(effective_formatter.get("priority"))
    effective_formatter["priority"] = priority

    row_spec = effective_formatter["row"]
    col_spec = effective_formatter["col"]

    # Base heading style (if header > 0)
    header_style = {
        "bold": True,
        "align": "center",
        "valign": "center",
        "text_wrap": True,
    }

    if apply_runtime_header_control:
        if priority == "col":
            _normalize_runtime_header_formats(
                col_spec,
                header_count=header_count,
                axis_size=num_cols,
                default_header_style=header_style,
            )
        else:
            _normalize_runtime_header_formats(
                row_spec,
                header_count=header_count,
                axis_size=num_rows,
                default_header_style=header_style,
            )

    _ensure_runtime_inferred_num_formats(
        effective_formatter,
        data_table=data_table,
        header_rows=header_count,
        priority=priority,
        num_rows=num_rows,
        num_cols=num_cols,
    )

    if engine == "xlsxwriter":
        _save_with_xlsxwriter(
            path=path,
            data_table=data_table,
            num_rows=num_rows,
            num_cols=num_cols,
            effective_formatter=effective_formatter,
            sheet_name=sheet_name,
            header_rows=header_count,
            progressor=progressor,
            as_template=as_template,
        )
    else:
        _save_with_openpyxl(
            path=path,
            data_table=data_table,
            num_rows=num_rows,
            num_cols=num_cols,
            effective_formatter=effective_formatter,
            sheet_name=sheet_name,
            header_rows=header_count,
            progressor=progressor,
            backup=backup,
            eff_mode=eff_mode,
            patch_skip_none=patch_skip_none,
            patch_skip_formulas=patch_skip_formulas,
            patch_skip_locked=patch_skip_locked,
            patch_range=patch_range,
            patch_strict_range=patch_strict_range,
            as_template=as_template,
        )

    return path


def save_formatted_xlsx(
        data_table: list[list[Any]],
        file_name: str | PathLike,
        formatter: SheetFormatSpec | None = None,
        *,
        header: int | None = None,
        sheet_name: str | None = None,
        mode: ModeLiteral = "auto",
        engine: EngineLiteral = "auto",
        backup: bool = True,
        patch_skip_none: bool = True,
        patch_skip_formulas: bool = True,
        patch_skip_locked: bool = True,
        patch_range: PatchRangeInput | None = None,
        patch_strict_range: bool = False,
        as_template: bool = False,
        progressor: PercentProgress | None = None,
) -> Path:
    """
    Backward-compatible alias for write_sheet(...).
    """
    return write_sheet(
        data_table=data_table,
        file_name=file_name,
        formatter=formatter,
        header=header,
        sheet_name=sheet_name,
        mode=mode,
        engine=engine,
        backup=backup,
        patch_skip_none=patch_skip_none,
        patch_skip_formulas=patch_skip_formulas,
        patch_skip_locked=patch_skip_locked,
        patch_range=patch_range,
        patch_strict_range=patch_strict_range,
        as_template=as_template,
        progressor=progressor,
    )


def _resolve_sheet_range_bounds(
        range_value: SheetRangeInput,
        *,
        param_name: str,
) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries

    try:
        if isinstance(range_value, str):
            min_col, min_row, max_col, max_row = range_boundaries(range_value)
        elif isinstance(range_value, tuple) and len(range_value) == 2:
            (start_row0, start_col0), (end_row0, end_col0) = range_value
            coords = (start_row0, start_col0, end_row0, end_col0)
            if any((not isinstance(v, int) or isinstance(v, bool) or v < 0) for v in coords):
                raise ValueError
            if end_row0 < start_row0 or end_col0 < start_col0:
                raise ValueError

            min_row = start_row0 + 1
            min_col = start_col0 + 1
            max_row = end_row0 + 1
            max_col = end_col0 + 1
        else:
            raise ValueError
    except Exception as e:
        raise ValueError(f"{inspect_name()}: Invalid {param_name}: {range_value!r}") from e

    return min_col, min_row, max_col, max_row


def read_sheet(
        file_name: str | PathLike,
        *,
        sheet_name: str | None = None,
        sheet_range: SheetRangeInput | None = None,
        data_only: bool = True,
) -> list[list[Any]]:
    """
    Reads worksheet values (all used cells or an explicit range) into list[list[Any]].
    """
    from openpyxl import load_workbook

    path = Path(file_name)
    if not path.exists():
        raise FileNotFoundError(f"{inspect_name()}: File not found: {path!s}")

    wb = load_workbook(str(path), data_only=data_only)
    ws = wb[sheet_name] if sheet_name is not None else wb.active

    if sheet_range is None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) == 1 and len(rows[0]) == 1 and rows[0][0] is None:
            return []
        return [list(row) for row in rows]

    min_col, min_row, max_col, max_row = _resolve_sheet_range_bounds(
        sheet_range,
        param_name="sheet_range",
    )
    rows = ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    return [list(row) for row in rows]


def _resolve_engine_mode_and_path(
        *,
        file_name: str | PathLike,
        engine: EngineLiteral,
        mode: ModeLiteral,
        as_template: bool,
) -> tuple[ResolvedEngineLiteral, ModeLiteral, Path]:
    """
    Brings engine/mode/path to a consistent state.
    """
    normalized_engine = cast(EngineLiteral, engine.lower())
    if normalized_engine not in ("auto", "xlsxwriter", "openpyxl"):
        raise ValueError(f"{inspect_name()}: Unsupported engine: {engine!r}")

    normalized_mode: ModeLiteral = mode.lower()  # type: ignore[assignment]
    if normalized_mode not in ("auto", "new", "replace", "update", "patch"):
        raise ValueError(f"{inspect_name()}: Unsupported mode: {mode!r}")

    file_name_with_ext = add_ext(file_name, ".xltx" if as_template else ".xlsx")
    path = Path(file_name_with_ext)
    file_exists = path.exists()

    if normalized_mode == "new" and file_exists:
        raise FileExistsError(
            f"{inspect_name()}: File already exists for mode 'new': {path!s}"
        )

    if normalized_mode == "auto":
        effective_mode: ModeLiteral = "replace" if file_exists else "new"
    else:
        effective_mode = normalized_mode

    if normalized_engine == "auto":
        resolved_engine: ResolvedEngineLiteral = (
            "openpyxl" if as_template or effective_mode != "new" else "xlsxwriter"
        )
    else:
        resolved_engine = cast(ResolvedEngineLiteral, normalized_engine)
        if normalized_mode == "auto":
            if file_exists and resolved_engine == "xlsxwriter":
                resolved_engine = "openpyxl"
                wrn(f'{inspect_name()}: Mode "auto": engine changed to "openpyxl", mode changed to "replace"')
            elif not file_exists and resolved_engine == "openpyxl" and not as_template:
                wrn(
                    f'{inspect_name()}: "Auto" mode for new file: Engine="openpyxl" specified. '
                    f'For new files, it is more efficient (speed, resources) to use engine="xlsxwriter".'
                )

    if resolved_engine == "xlsxwriter" and effective_mode != "new":
        raise ValueError(
            f"{inspect_name()}: Mode {effective_mode!r} is not supported for engine 'xlsxwriter'"
        )

    return resolved_engine, effective_mode, path


def _get_style_formatter(
        effective_formatter: SheetFormatSpec,
        *,
        num_rows: int,
        num_cols: int,
        engine: ResolvedEngineLiteral,
        workbook: Any = None,
) -> dict[int, list[Any]]:
    """
    Builds a native formatter or returns an empty dictionary if there is nothing to format.
    """
    if num_cols <= 0 or num_rows <= 0:
        return {}

    logical_formatter: RowFormats = build_row_formats_with_columns(
        effective_formatter,
        max_cols=num_cols,
        max_rows=num_rows,
        inherit_defaults=True,
    )
    return translate_formatter(
        logical_formatter,
        engine=engine,
        workbook=workbook,
    )


def _get_row_styles(style_formatter: dict[int, list[Any]], row_idx: int) -> list[Any] | None:
    """
    Returns a list of styles for a string, taking into account explicit formatting and zebra.
    """
    if row_idx in style_formatter:
        return style_formatter[row_idx]
    if -2 in style_formatter:
        zebra_start = 0
        while zebra_start in style_formatter:
            zebra_start += 1
        template_idx = -1 if ((row_idx - zebra_start) % 2 == 0) else -2
        if template_idx in style_formatter:
            return style_formatter[template_idx]
    return style_formatter.get(-1)


def _save_with_xlsxwriter(
    *,
    path: Path,
    data_table: list[list[Any]],
    num_rows: int,
    num_cols: int,
    effective_formatter: SheetFormatSpec,
    sheet_name: str | None,
    header_rows: int,
    progressor: PercentProgress | None,
    as_template: bool,
    ) -> None:
    """
    Saves file via xlsxwriter.
    """
    import xlsxwriter

    if as_template:
        raise ValueError(
            f"{inspect_name()}: xlsxwriter does not support saving Excel templates (.xltx). "
            f"To save as template, use engine='openpyxl'."
        )

    workbook_options = {
        "constant_memory": bool(TRY_OPTIMIZE),
        "strings_to_urls": False,
    }
    workbook = xlsxwriter.Workbook(str(path), workbook_options)
    worksheet = workbook.add_worksheet(sheet_name) if sheet_name is not None else workbook.add_worksheet()

    style_formatter = _get_style_formatter(
        effective_formatter,
        num_rows=num_rows,
        num_cols=num_cols,
        engine="xlsxwriter",
        workbook=workbook,
    )

    for r_idx, row in enumerate(data_table):
        row_styles = _get_row_styles(style_formatter, r_idx)
        for c_idx, value in enumerate(row):
            if row_styles and c_idx < len(row_styles):
                worksheet.write(r_idx, c_idx, value, row_styles[c_idx])
            else:
                worksheet.write(r_idx, c_idx, value)

        if progressor is not None:
            progressor.update(r_idx)

    # Deterministic width estimation is more stable than engine autofit()
    # for wrapped headers and multi-language content.
    col_widths = compute_column_widths(
        data_table,
        header_rows=header_rows,
        max_width=MAX_COLUMN_WIDTH['chars'],
        reserve_chars=COLUMN_WIDTH_RESERVE_CHARS,
    )
    for col, width in enumerate(col_widths):
        worksheet.set_column(col, col, width)

    workbook.close()


def _save_with_openpyxl(
    *,
    path: Path,
    data_table: list[list[Any]],
    num_rows: int,
    num_cols: int,
    effective_formatter: SheetFormatSpec,
    sheet_name: str | None,
    header_rows: int,
    progressor: PercentProgress | None,
    backup: bool,
    eff_mode: str,
    patch_skip_none: bool,
    patch_skip_formulas: bool,
    patch_skip_locked: bool,
    patch_range: PatchRangeInput | None,
    patch_strict_range: bool,
    as_template: bool,
    ) -> None:
    """
   Saves the file using openpyxl.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter

    file_exists = path.exists()

    if eff_mode == "new" or not file_exists:
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name
        if as_template:
            wb.template = True
    else:
        wb = load_workbook(str(path))
        target_name = sheet_name or wb.active.title

        if eff_mode == "replace":
            if target_name in wb.sheetnames:
                sheet_index = wb.worksheets.index(wb[target_name])
                if backup:
                    backup_name = rotate_sheet_versions(wb, target_name)
                    if isinstance(backup_name, str):
                        retarget_sheet_references(
                            wb,
                            old_name=backup_name,
                            new_name=target_name,
                            exclude_sheets={backup_name},
                        )
                    ws = wb.create_sheet(title=target_name, index=sheet_index)
                else:
                    ws_old = wb[target_name]
                    sheet_index = wb.worksheets.index(ws_old)
                    wb.remove(ws_old)
                    ws = wb.create_sheet(title=target_name, index=sheet_index)
            else:
                ws = wb.create_sheet(title=target_name, index=0)

        elif eff_mode == "update":
            if target_name in wb.sheetnames:
                ws = wb[target_name]
                if backup:
                    backup_name = rotate_sheet_versions(wb, target_name, copy_sheet=True)
                    if isinstance(backup_name, str):
                        retarget_sheet_references(
                            wb,
                            old_name=backup_name,
                            new_name=target_name,
                            exclude_sheets={backup_name},
                        )

                max_row = ws.max_row
                max_col = ws.max_column
                clear_range_openpyxl(max_col, max_row, ws)
            else:
                ws = wb.create_sheet(title=target_name)
        elif eff_mode == "patch":
            if not file_exists:
                raise FileNotFoundError(f"{inspect_name()}: mode 'patch' requires an existing file: {path!s}")
            if target_name not in wb.sheetnames:
                raise ValueError(f"{inspect_name()}: Sheet {target_name!r} not found for mode 'patch'")
            ws = wb[target_name]
        else:
            raise ValueError(f"{inspect_name()}: Unsupported effective mode for openpyxl: {eff_mode!r}")

    if eff_mode == "patch":
        _patch_sheet_values_openpyxl(
            ws,
            data_table=data_table,
            progressor=progressor,
            patch_skip_none=patch_skip_none,
            patch_skip_formulas=patch_skip_formulas,
            patch_skip_locked=patch_skip_locked,
            patch_range=patch_range,
            patch_strict_range=patch_strict_range,
        )
        wb.save(str(path))
        return

    style_formatter = _get_style_formatter(
        effective_formatter,
        num_rows=num_rows,
        num_cols=num_cols,
        engine="openpyxl",
    )

    for r_idx, row in enumerate(data_table):
        excel_row = r_idx + 1
        row_styles = _get_row_styles(style_formatter, r_idx)

        for c_idx, value in enumerate(row):
            cell = ws.cell(row=excel_row, column=c_idx + 1, value=value)
            if row_styles and c_idx < len(row_styles):
                apply_style(cell, row_styles[c_idx])

        if progressor is not None:
            progressor.update(r_idx)

    col_widths = compute_column_widths(
        data_table,
        header_rows=header_rows,
        reserve_chars=COLUMN_WIDTH_RESERVE_CHARS,
    )
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(str(path))


def _patch_sheet_values_openpyxl(
        ws: Any,
        *,
        data_table: list[list[Any]],
        progressor: PercentProgress | None,
        patch_skip_none: bool,
        patch_skip_formulas: bool,
        patch_skip_locked: bool,
        patch_range: PatchRangeInput | None,
        patch_strict_range: bool,
) -> None:
    """
    Applies sparse value updates to an existing worksheet while preserving layout and styles.
    """
    start_row = 1
    start_col = 1
    max_row_limit: int | None = None
    max_col_limit: int | None = None

    if patch_range is not None:
        min_col, min_row, max_col, max_row = _resolve_sheet_range_bounds(
            patch_range,
            param_name="patch_range",
        )
        start_row = min_row
        start_col = min_col
        max_row_limit = max_row
        max_col_limit = max_col

    max_data_cols = max((len(row) for row in data_table), default=0)
    if patch_range is not None and patch_strict_range:
        allowed_rows = max(0, (max_row_limit or start_row) - start_row + 1)
        allowed_cols = max(0, (max_col_limit or start_col) - start_col + 1)
        if len(data_table) > allowed_rows or max_data_cols > allowed_cols:
            raise ValueError(
                f"{inspect_name()}: data_table shape ({len(data_table)}x{max_data_cols}) "
                f"does not fit patch_range {patch_range!r} ({allowed_rows}x{allowed_cols})"
            )

    sheet_protected = bool(getattr(getattr(ws, "protection", None), "sheet", False))

    for r_idx, row in enumerate(data_table):
        excel_row = start_row + r_idx
        if max_row_limit is not None and excel_row > max_row_limit:
            break
        for c_idx, value in enumerate(row):
            excel_col = start_col + c_idx
            if max_col_limit is not None and excel_col > max_col_limit:
                break
            if patch_skip_none and value is None:
                continue

            cell = ws.cell(row=excel_row, column=excel_col)

            if patch_skip_formulas and isinstance(cell.value, str) and cell.value.startswith("="):
                continue

            if patch_skip_locked and sheet_protected:
                locked = getattr(getattr(cell, "protection", None), "locked", False)
                if bool(locked):
                    continue

            cell.value = value

        if progressor is not None:
            progressor.update(r_idx)
