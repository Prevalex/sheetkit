from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from sheetkit import extract_formatter_to_file

EXAMPLE_DIR = Path(__file__).resolve().parent
TEMPLATE_FILE = EXAMPLE_DIR / "07_formatter_template.xlsx"
PRESET_FILE = EXAMPLE_DIR / "07_imported_formatter.json"


def build_template_workbook(file_name: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "StyledTable"

    header1_fill = PatternFill(patternType="solid", fgColor="4472C4")
    header2_fill = PatternFill(patternType="solid", fgColor="D9E2F3")
    alt_fill = PatternFill(patternType="solid", fgColor="F7F9FC")
    white_font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    dark_font = Font(name="Aptos", size=11, bold=True, color="1F1F1F")
    data_font = Font(name="Aptos", size=11, color="222222")
    blue_line = Side(style="thin", color="D9E2F3")
    gray_line = Side(style="thin", color="D9D9D9")

    for col_idx in range(1, 5):
        ws.cell(row=1, column=col_idx, value=f"Header 1 / {col_idx}")
        ws.cell(row=1, column=col_idx).fill = header1_fill
        ws.cell(row=1, column=col_idx).font = white_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=col_idx).border = Border(bottom=blue_line)

        ws.cell(row=2, column=col_idx, value=f"Header 2 / {col_idx}")
        ws.cell(row=2, column=col_idx).fill = header2_fill
        ws.cell(row=2, column=col_idx).font = dark_font
        ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col_idx).border = Border(bottom=blue_line)

        ws.cell(row=3, column=col_idx, value=10 * col_idx)
        ws.cell(row=3, column=col_idx).font = data_font
        ws.cell(row=3, column=col_idx).border = Border(bottom=gray_line)

        ws.cell(row=4, column=col_idx, value=20 * col_idx)
        ws.cell(row=4, column=col_idx).font = data_font
        ws.cell(row=4, column=col_idx).fill = alt_fill
        ws.cell(row=4, column=col_idx).border = Border(bottom=gray_line)

    ws["A3"].number_format = "@"
    ws["B3"].number_format = "#,##0.00"
    ws["C3"].number_format = "0.0%"
    ws["D4"].number_format = "yyyy-mm-dd"

    wb.save(file_name)


build_template_workbook(TEMPLATE_FILE)

path = extract_formatter_to_file(
    file_name=TEMPLATE_FILE,
    json_file=PRESET_FILE,
    columns=4,
    start_cell="A1",
    header=2,
    zebra=True,
    name="excel_drawn_blue",
)

print(path.resolve())

